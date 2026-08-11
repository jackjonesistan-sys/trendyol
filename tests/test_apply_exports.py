import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

import app


def write_workbook(path, headers, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class ApplyExportTests(unittest.TestCase):
    def setUp(self):
        app.app.config["TESTING"] = True
        self.client = app.app.test_client()

    def apply_from_temp(self, root, result_rows, input_files, payload):
        output_dir = root / "output"
        output_dir.mkdir()
        result_path = root / "result.xlsx"
        pd.DataFrame(result_rows).to_excel(result_path, index=False)
        manifest_path = root / "manifest.json"
        if not manifest_path.exists():
            manifest_path.write_text('{"files": {}}', encoding="utf-8")

        with (
            patch.object(app, "OUTPUT_DIR", str(output_dir)),
            patch.object(app, "F_HESAP", str(result_path)),
            patch.object(app, "INPUT_MANIFEST", str(manifest_path)),
            patch.object(app, "load_upload_set", return_value=input_files),
            patch.object(app, "fix_xlsx_for_trendyol", side_effect=lambda _path: None),
            patch(
                "fiyat_farki_analiz_script.generate_fiyat_farki_raporu",
                side_effect=lambda *_args, **_kwargs: None,
            ),
        ):
            response = self.client.post("/api/apply", json=payload)
        return response, output_dir

    def test_accounting_only_rows_are_exported_for_all_main_campaigns(self):
        cases = (
            {
                "campaign": "Avantajlı",
                "input_key": "muhasebe_avantaj",
                "headers": ["BARKOD", "YENİ TSF (FİYAT GÜNCELLE)"],
                "source_row": ["A-1", 80],
                "result_price": {"Avantajlı Ürün Fiyatı (YENİ TSF) (TL)": 75},
                "filename": "Avantajli_Urun_Etiketleri.xlsx",
                "price_header": "YENİ TSF (FİYAT GÜNCELLE)",
                "price": 75,
            },
            {
                "campaign": "Flaş",
                "input_key": "muhasebe_flas",
                "headers": ["Barkod", "Senin Belirlediğin Flaş Fiyatı"],
                "source_row": ["F-1", 85],
                "result_price": {"Flaş Ürün 24 Saat Fiyatı (TL)": 82},
                "filename": "Flas_Urunler_Genel.xlsx",
                "price_header": "24 Saat Fiyat",
                "price": 82,
            },
            {
                "campaign": "Plus",
                "input_key": "muhasebe_plus",
                "headers": ["Barkod", "Plus Fiyat Üst Limiti"],
                "source_row": ["P-1", 90],
                "result_price": {"Plus Fiyatı (TL)": 88},
                "filename": "Plus_Komisyon_Tarifeleri.xlsx",
                "price_header": "Plus Fiyat Seçimi",
                "price": 88,
            },
        )

        for case in cases:
            with self.subTest(campaign=case["campaign"]), tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                source = root / f'{case["input_key"]}.xlsx'
                write_workbook(source, case["headers"], [case["source_row"]])
                barcode = case["source_row"][0]
                row = {
                    "Barkod": barcode,
                    "Stok Adedi": 1,
                    "İlk Kampanya Seçimi": "Hiçbiri",
                    "İlk Ekstra Kampanya Seçimi": "Hiçbiri",
                    "Uygulanabilir Kampanyalar": case["campaign"],
                    "eligible_main_campaigns": ["Hiçbiri", case["campaign"]],
                    "eligible_extra_campaigns": ["Hiçbiri"],
                    **case["result_price"],
                }

                response, output_dir = self.apply_from_temp(
                    root,
                    [row],
                    {case["input_key"]: str(source)},
                    {
                        "target_type": case["campaign"],
                        "selections": {
                            barcode: {"main": case["campaign"], "extra": "Hiçbiri"}
                        },
                    },
                )

                self.assertEqual(response.status_code, 200, response.get_json())
                run_dir = output_dir / response.get_json()["timestamp_folder"]
                exported = openpyxl.load_workbook(run_dir / case["filename"]).active
                headers = [cell.value for cell in exported[1]]
                barcode_column = next(
                    index for index, value in enumerate(headers, 1)
                    if str(value).strip().casefold() == "barkod"
                )
                price_column = headers.index(case["price_header"]) + 1
                self.assertEqual(exported.cell(2, barcode_column).value, barcode)
                self.assertEqual(exported.cell(2, price_column).value, case["price"])
                if case["campaign"] == "Plus":
                    self.assertEqual(
                        exported.cell(2, headers.index("Tarife Seçimi") + 1).value,
                        "7 Günlük Fiyat",
                    )

    def test_flash_export_uses_each_interval_evaluation_price(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            headers = [
                "Barkod",
                "Güncellenecek Fiyat",
                "24 Saat Fiyat",
                "Senin Belirlediğin Flaş Fiyatı",
                "24 Saat Flaş Başlangıç Tarihi",
                "24 Saat Flaş Bitiş Tarihi",
            ]
            write_workbook(
                source,
                headers,
                [
                    ["ST34-4070", None, 968.81, None, "10/08/2026 00:00", "10/08/2026 23:59"],
                    ["ST34-4070", None, 960.64, None, "12/08/2026 00:00", "12/08/2026 23:59"],
                ],
            )
            row = {
                "Barkod": "ST34-4070",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "Flaş Ürün 24 Saat Fiyatı (TL)": 968.81,
                "flash_evaluations": [
                    {
                        "period": "24 Saat",
                        "start": "2026-08-10 00:00:00",
                        "end": "2026-08-10 23:59:00",
                        "price": 968.81,
                        "rate": 19.1,
                        "net": 783.77,
                        "eligible": True,
                        "source": "24 Saat Fiyat",
                    },
                    {
                        "period": "24 Saat",
                        "start": "12/08/2026 00:00",
                        "end": "12/08/2026 23:59",
                        "price": 960.64,
                        "rate": 19.1,
                        "net": 777.16,
                        "eligible": True,
                        "source": "24 Saat Fiyat",
                    },
                ],
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {
                        "ST34-4070": {"main": "Flaş", "extra": "Hiçbiri"}
                    },
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            run_dir = output_dir / response.get_json()["timestamp_folder"]
            expected = {
                "Flas_Urunler_10_08_2026.xlsx": 968.81,
                "Flas_Urunler_12_08_2026.xlsx": 960.64,
            }
            flash_names = {
                Path(name).name
                for name in response.get_json()["generated_files"]
                if Path(name).name.startswith("Flas_Urunler_")
            }
            self.assertEqual(flash_names, set(expected))
            for filename, price in expected.items():
                exported = openpyxl.load_workbook(run_dir / filename).active
                output_headers = [cell.value for cell in exported[1]]
                self.assertEqual(exported.max_row, 2)
                self.assertEqual(
                    exported.cell(2, output_headers.index("24 Saat Fiyat") + 1).value,
                    price,
                )
                self.assertEqual(
                    exported.cell(2, output_headers.index("Güncellenecek Fiyat") + 1).value,
                    "24 Saat",
                )

    def test_flash_export_omits_ineligible_interval_and_applies_fixed_price_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            headers = [
                "Barkod",
                "Güncellenecek Fiyat",
                "24 Saat Fiyat",
                "Senin Belirlediğin Flaş Fiyatı",
                "24 Saat Flaş Başlangıç Tarihi",
                "24 Saat Flaş Bitiş Tarihi",
            ]
            write_workbook(
                source,
                headers,
                [
                    ["FIXED", None, 999, 950.25, "10/08/2026 00:00", "10/08/2026 23:59"],
                    ["FIXED", None, 930, 920, "12/08/2026 00:00", "12/08/2026 23:59"],
                ],
            )
            row = {
                "Barkod": "FIXED",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "Flaş Ürün 24 Saat Fiyatı (TL)": 999,
                "flash_evaluations": [
                    {
                        "period": "24 Saat",
                        "start": "10/08/2026 00:00",
                        "end": "10/08/2026 23:59",
                        "price": 950.25,
                        "rate": 19.1,
                        "net": 768.75,
                        "eligible": True,
                        "source": "Senin Belirlediğin Flaş Fiyatı",
                    },
                    {
                        "period": "24 Saat",
                        "start": "12/08/2026 00:00",
                        "end": "12/08/2026 23:59",
                        "price": 920,
                        "rate": 19.1,
                        "net": 744.28,
                        "eligible": False,
                        "source": "Senin Belirlediğin Flaş Fiyatı",
                    },
                ],
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {"FIXED": {"main": "Flaş", "extra": "Hiçbiri"}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            flash_names = [
                Path(name).name
                for name in response.get_json()["generated_files"]
                if Path(name).name.startswith("Flas_Urunler_")
            ]
            self.assertEqual(flash_names, ["Flas_Urunler_10_08_2026.xlsx"])
            exported = openpyxl.load_workbook(
                output_dir / response.get_json()["timestamp_folder"] / flash_names[0]
            ).active
            output_headers = [cell.value for cell in exported[1]]
            self.assertEqual(
                exported.cell(
                    2,
                    output_headers.index("Senin Belirlediğin Flaş Fiyatı") + 1,
                ).value,
                950.25,
            )
            self.assertEqual(
                exported.cell(2, output_headers.index("Güncellenecek Fiyat") + 1).value,
                "Senin Belirlediğin Flaş Fiyatı",
            )

    def test_flash_export_binds_single_undated_accounting_fixed_price_to_template_interval(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            standard = root / "flash.xlsx"
            accounting = root / "accounting.xlsx"
            write_workbook(
                standard,
                [
                    "Barkod",
                    "Güncellenecek Fiyat",
                    "24 Saat Fiyat",
                    "24 Saat Flaş Başlangıç Tarihi",
                    "24 Saat Flaş Bitiş Tarihi",
                ],
                [["FIXED", None, 999, "10/08/2026 00:00", "10/08/2026 23:59"]],
            )
            write_workbook(
                accounting,
                ["Barkod", "Senin Belirlediğin Flaş Fiyatı"],
                [["FIXED", 950.25]],
            )
            row = {
                "Barkod": "FIXED",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "flash_evaluations": [{
                    "period": "24 Saat",
                    "start": None,
                    "end": None,
                    "price": 950.25,
                    "rate": 19.1,
                    "net": 768.75,
                    "eligible": True,
                    "source": "Senin Belirlediğin Flaş Fiyatı",
                }],
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(standard), "muhasebe_flas": str(accounting)},
                {
                    "target_type": "Flaş",
                    "selections": {"FIXED": {"main": "Flaş", "extra": "Hiçbiri"}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            filename = "Flas_Urunler_10_08_2026.xlsx"
            generated = {Path(name).name for name in response.get_json()["generated_files"]}
            self.assertIn(filename, generated)
            exported = openpyxl.load_workbook(
                output_dir / response.get_json()["timestamp_folder"] / filename
            ).active
            output_headers = [cell.value for cell in exported[1]]
            self.assertEqual(
                exported.cell(
                    2,
                    output_headers.index("Senin Belirlediğin Flaş Fiyatı") + 1,
                ).value,
                950.25,
            )
            self.assertEqual(
                exported.cell(2, output_headers.index("Güncellenecek Fiyat") + 1).value,
                "Senin Belirlediğin Flaş Fiyatı",
            )

    def test_flash_export_rejects_undated_fixed_evaluation_for_multiple_intervals(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            write_workbook(
                source,
                [
                    "Barkod",
                    "Güncellenecek Fiyat",
                    "24 Saat Fiyat",
                    "24 Saat Flaş Başlangıç Tarihi",
                    "24 Saat Flaş Bitiş Tarihi",
                ],
                [
                    ["FIXED", None, 999, "10/08/2026 00:00", "10/08/2026 23:59"],
                    ["FIXED", None, 990, "12/08/2026 00:00", "12/08/2026 23:59"],
                ],
            )
            row = {
                "Barkod": "FIXED",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "flash_evaluations": [{
                    "period": "24 Saat",
                    "start": None,
                    "end": None,
                    "price": 950.25,
                    "rate": 19.1,
                    "net": 768.75,
                    "eligible": True,
                    "source": "Senin Belirlediğin Flaş Fiyatı",
                }],
            }

            response, _output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {"FIXED": {"main": "Flaş", "extra": "Hiçbiri"}},
                },
            )

            self.assertEqual(response.status_code, 400, response.get_json())
            self.assertIn("yeniden hesaplayın", response.get_json()["message"])

    def test_flash_legacy_scalar_is_not_repeated_across_duplicate_intervals(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            write_workbook(
                source,
                [
                    "Barkod",
                    "Güncellenecek Fiyat",
                    "24 Saat Fiyat",
                    "24 Saat Flaş Başlangıç Tarihi",
                    "24 Saat Flaş Bitiş Tarihi",
                ],
                [
                    ["LEGACY", None, 968.81, "10/08/2026 00:00", "10/08/2026 23:59"],
                    ["LEGACY", None, 960.64, "12/08/2026 00:00", "12/08/2026 23:59"],
                ],
            )
            row = {
                "Barkod": "LEGACY",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "Flaş Ürün 24 Saat Fiyatı (TL)": 968.81,
            }

            response, _output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {
                        "LEGACY": {"main": "Flaş", "extra": "Hiçbiri"}
                    },
                },
            )

            self.assertEqual(response.status_code, 400, response.get_json())
            self.assertIn("yeniden hesaplayın", response.get_json()["message"])

    def test_flash_export_separates_24_and_3_hour_intervals_on_same_date(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            write_workbook(
                source,
                [
                    "Barkod",
                    "Güncellenecek Fiyat",
                    "24 Saat Fiyat",
                    "3 Saat Fiyat",
                    "24 Saat Flaş Başlangıç Tarihi",
                    "24 Saat Flaş Bitiş Tarihi",
                    "3 Saat Flaş Başlangıç Tarihi",
                    "3 Saat Flaş Bitiş Tarihi",
                ],
                [[
                    "BOTH",
                    None,
                    900,
                    925,
                    "13/08/2026 00:00",
                    "13/08/2026 23:59",
                    "13/08/2026 10:00",
                    "13/08/2026 12:59",
                ]],
            )
            row = {
                "Barkod": "BOTH",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "flash_evaluations": [
                    {
                        "period": "24 Saat",
                        "start": "13/08/2026 00:00",
                        "end": "13/08/2026 23:59",
                        "price": 900,
                        "rate": 19,
                        "net": 729,
                        "eligible": True,
                        "source": "24 Saat Fiyat",
                    },
                    {
                        "period": "3 Saat",
                        "start": "13/08/2026 10:00",
                        "end": "13/08/2026 12:59",
                        "price": 925,
                        "rate": 18,
                        "net": 758.5,
                        "eligible": True,
                        "source": "3 Saat Fiyat",
                    },
                ],
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {"BOTH": {"main": "Flaş", "extra": "Hiçbiri"}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            run_dir = output_dir / response.get_json()["timestamp_folder"]
            expected = {
                "Flas_Urunler_13_08_2026.xlsx": ("24 Saat Fiyat", 900, "24 Saat"),
                "Flas_Urunler_13_08_2026_3_Saat.xlsx": ("3 Saat Fiyat", 925, "3 Saat"),
            }
            names = {
                Path(name).name
                for name in response.get_json()["generated_files"]
                if Path(name).name.startswith("Flas_Urunler_")
            }
            self.assertEqual(names, set(expected))
            for filename, (price_header, price, selection) in expected.items():
                exported = openpyxl.load_workbook(run_dir / filename).active
                output_headers = [cell.value for cell in exported[1]]
                self.assertEqual(
                    exported.cell(2, output_headers.index(price_header) + 1).value,
                    price,
                )
                self.assertEqual(
                    exported.cell(2, output_headers.index("Güncellenecek Fiyat") + 1).value,
                    selection,
                )

    def test_flash_export_keeps_price_only_three_hour_interval(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "flash.xlsx"
            write_workbook(
                source,
                ["Barkod", "Güncellenecek Fiyat", "24 Saat Fiyat", "3 Saat Fiyat"],
                [["THREE", None, None, 925]],
            )
            row = {
                "Barkod": "THREE",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": "Flaş",
                "eligible_main_campaigns": ["Hiçbiri", "Flaş"],
                "eligible_extra_campaigns": ["Hiçbiri"],
                "flash_evaluations": [{
                    "period": "3 Saat",
                    "start": None,
                    "end": None,
                    "price": 925,
                    "rate": 18,
                    "net": 758.5,
                    "eligible": True,
                    "source": "3 Saat Fiyat",
                }],
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {"flash": str(source)},
                {
                    "target_type": "Flaş",
                    "selections": {"THREE": {"main": "Flaş", "extra": "Hiçbiri"}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            filename = "Flas_Urunler_Genel_3_Saat.xlsx"
            generated = {Path(name).name for name in response.get_json()["generated_files"]}
            self.assertIn(filename, generated)
            exported = openpyxl.load_workbook(
                output_dir / response.get_json()["timestamp_folder"] / filename
            ).active
            output_headers = [cell.value for cell in exported[1]]
            self.assertEqual(
                exported.cell(2, output_headers.index("3 Saat Fiyat") + 1).value,
                925,
            )
            self.assertEqual(
                exported.cell(2, output_headers.index("Güncellenecek Fiyat") + 1).value,
                "3 Saat",
            )

    def test_plus_multi_period_export_groups_all_intervals_and_preserves_template(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "plus.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = [
                "Barkod",
                "Tarih Aralığı (2 Gün)",
                "Plus Komisyon Teklifi",
                "Hesaplanan Komisyon (2 Gün)",
                "Tarih Aralığı (5 Gün)",
                "Plus Komisyon Teklifi",
                "Hesaplanan Komisyon (5 Gün)",
                "Plus Fiyat Üst Limiti",
                "Plus Fiyat Seçimi",
                "Tarife Seçimi",
                "İptal",
                "2 Gün Tarih Aralığı",
                "5 Gün Tarih Aralığı",
                "7 Gün Tarih Aralığı",
            ]
            sheet.append(headers)
            interval_2 = "11.08.2026 - 12.08.2026"
            interval_5 = "20.08.2026 - 24.08.2026"
            label_2 = "2 Günlük Fiyat (11 Ağustos 08.00-13 Ağustos 07.59)"
            label_5 = "5 Günlük Fiyat (20 Ağustos 08.00-25 Ağustos 07.59)"
            source_rows = (
                (
                    "BOTH", interval_2, 12, None, interval_5, 15, None, 99,
                    None, None, None, label_2, label_5, "7 Günlük Fiyat",
                ),
                (
                    "FIRST", interval_2, 13, None, None, None, None, 89,
                    None, None, None, label_2, None, None,
                ),
                (
                    "SECOND", None, None, None, interval_5, 16, None, 79,
                    None, None, None, label_5, None, None,
                ),
            )
            for row_number, values in enumerate(source_rows, 2):
                sheet.append(values)
                sheet.cell(row_number, 4).value = f'=IF(C{row_number}="","-",C{row_number}*2)'
                sheet.cell(row_number, 7).value = f'=IF(F{row_number}="","-",F{row_number}*2)'
                sheet.cell(row_number, 11).value = f'=IF(ISBLANK(I{row_number}),"","Hayır")'
                sheet.cell(row_number, 4).number_format = "0.00"
                sheet.cell(row_number, 7).number_format = "0.00"
            validation = DataValidation(
                type="list",
                formula1="OFFSET($L2,0,0,1,COUNTA($L2:$N2))",
            )
            sheet.add_data_validation(validation)
            validation.add("J2:J4")
            sheet.add_table(Table(displayName="PlusTarifeleri", ref="A1:N4"))
            workbook.save(source)

            tariffs = {
                "BOTH": (91, "7 Günlük Fiyat"),
                "FIRST": (81, label_2),
                "SECOND": (71, label_5),
            }
            rows = [
                {
                    "Barkod": barcode,
                    "Stok Adedi": 1,
                    "Uygulanabilir Kampanyalar": "Plus",
                    "eligible_main_campaigns": ["Hiçbiri", "Plus"],
                    "eligible_extra_campaigns": ["Hiçbiri"],
                    "Plus Fiyatı (TL)": price,
                    "Plus Tarife Seçimi": tariff,
                }
                for barcode, (price, tariff) in tariffs.items()
            ]

            response, output_dir = self.apply_from_temp(
                root,
                rows,
                {"plus": str(source)},
                {
                    "target_type": "Plus",
                    "selections": {
                        barcode: {"main": "Plus", "extra": "Hiçbiri"}
                        for barcode in tariffs
                    },
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            plus_files = [
                output_dir / name
                for name in response.get_json()["generated_files"]
                if Path(name).name.startswith("Plus_Komisyon_Tarifeleri")
            ]
            self.assertEqual(len(plus_files), 3)
            exports = {}
            for path in plus_files:
                exported = openpyxl.load_workbook(path, data_only=False).active
                self.assertEqual(exported.max_row, 2)
                barcode = exported["A2"].value
                exports[barcode] = exported
                self.assertEqual(
                    [cell.value for cell in exported[1]].count("Plus Komisyon Teklifi"),
                    2,
                )
                self.assertEqual(exported["D2"].value, '=IF(C2="","-",C2*2)')
                self.assertEqual(exported["G2"].value, '=IF(F2="","-",F2*2)')
                self.assertEqual(exported["K2"].value, '=IF(ISBLANK(I2),"","Hayır")')
                self.assertEqual(exported["D2"].number_format, "0.00")
                self.assertEqual(exported["G2"].number_format, "0.00")
                self.assertIn("PlusTarifeleri", exported.tables)
                self.assertEqual(exported.tables["PlusTarifeleri"].ref, "A1:N2")
                self.assertEqual(len(exported.data_validations.dataValidation), 1)
                output_validation = exported.data_validations.dataValidation[0]
                self.assertEqual(str(output_validation.sqref), "J2")
                self.assertEqual(output_validation.formula1, validation.formula1)

            self.assertEqual(set(exports), set(tariffs))
            for barcode, (price, tariff) in tariffs.items():
                self.assertEqual(exports[barcode]["I2"].value, price)
                self.assertEqual(exports[barcode]["J2"].value, tariff)

    def test_plus_legacy_results_use_detected_day_sum_and_price_fallback(self):
        cases = (
            (
                ["Tarih Aralığı (9 Gün)", "Plus Komisyon Teklifi"],
                ["11.08.2026 - 19.08.2026", 12],
                88,
                "9 Günlük Fiyat",
                "Plus_Komisyon_Tarifeleri_11.08.2026_-_19.08.2026.xlsx",
            ),
            (
                [
                    "Tarih Aralığı (2 Gün)",
                    "Plus Komisyon Teklifi",
                    "Tarih Aralığı (5 Gün)",
                    "Plus Komisyon Teklifi",
                ],
                ["11.08.2026 - 12.08.2026", 12, "20.08.2026 - 24.08.2026", 15],
                None,
                "7 Günlük Fiyat",
                None,
            ),
        )
        for period_headers, period_values, result_price, expected_tariff, expected_name in cases:
            with self.subTest(tariff=expected_tariff), tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                source = root / "plus.xlsx"
                write_workbook(
                    source,
                    [
                        "Barkod",
                        *period_headers,
                        "Plus Fiyat Üst Limiti",
                        "Plus Fiyat Seçimi",
                        "Tarife Seçimi",
                    ],
                    [["LEGACY", *period_values, 90, None, None]],
                )
                row = {
                    "Barkod": "LEGACY",
                    "Stok Adedi": 1,
                    "Uygulanabilir Kampanyalar": "Plus",
                    "eligible_main_campaigns": ["Hiçbiri", "Plus"],
                    "eligible_extra_campaigns": ["Hiçbiri"],
                }
                if result_price is not None:
                    row["Plus Fiyatı (TL)"] = result_price

                response, output_dir = self.apply_from_temp(
                    root,
                    [row],
                    {"plus": str(source)},
                    {
                        "target_type": "Plus",
                        "selections": {"LEGACY": {"main": "Plus", "extra": "Hiçbiri"}},
                    },
                )

                self.assertEqual(response.status_code, 200, response.get_json())
                output_name = next(
                    name for name in response.get_json()["generated_files"]
                    if Path(name).name.startswith("Plus_Komisyon_Tarifeleri")
                )
                if expected_name:
                    self.assertEqual(Path(output_name).name, expected_name)
                exported = openpyxl.load_workbook(output_dir / output_name).active
                exported_headers = [cell.value for cell in exported[1]]
                self.assertEqual(
                    exported.cell(2, exported_headers.index("Plus Fiyat Seçimi") + 1).value,
                    result_price if result_price is not None else 90,
                )
                self.assertEqual(
                    exported.cell(2, exported_headers.index("Tarife Seçimi") + 1).value,
                    expected_tariff,
                )

    def test_standard_template_wins_duplicates_and_accounting_only_row_is_appended(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            standard = root / "advantage.xlsx"
            accounting = root / "accounting.xlsx"
            headers = [
                "BARKOD",
                "YENİ TSF (FİYAT GÜNCELLE)",
                "Tarife Sonuna Kadar Uygula",
                "Kaynak",
            ]
            write_workbook(standard, headers, [["DUP", 90, None, "standart"]])
            write_workbook(
                accounting,
                headers,
                [["DUP", 70, None, "muhasebe"], ["ONLY", 75, None, "muhasebe"]],
            )
            rows = [
                {
                    "Barkod": barcode,
                    "Stok Adedi": 1,
                    "Uygulanabilir Kampanyalar": "Avantajlı",
                    "eligible_main_campaigns": ["Hiçbiri", "Avantajlı"],
                    "eligible_extra_campaigns": ["Hiçbiri"],
                    "Avantajlı Ürün Fiyatı (YENİ TSF) (TL)": price,
                }
                for barcode, price in (("DUP", 81), ("ONLY", 72))
            ]

            response, output_dir = self.apply_from_temp(
                root,
                rows,
                {
                    "advantage": str(standard),
                    "muhasebe_avantaj": str(accounting),
                },
                {
                    "target_type": "Avantajlı",
                    "selections": {
                        barcode: {"main": "Avantajlı", "extra": "Hiçbiri"}
                        for barcode in ("DUP", "ONLY")
                    },
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            run_dir = output_dir / response.get_json()["timestamp_folder"]
            sheet = openpyxl.load_workbook(run_dir / "Avantajli_Urun_Etiketleri.xlsx").active
            exported = {
                sheet.cell(row, 1).value: sheet.cell(row, 4).value
                for row in range(2, sheet.max_row + 1)
            }
            self.assertEqual(exported, {"DUP": "standart", "ONLY": "muhasebe"})

    def test_dict_selection_preserves_extra_export_and_excludes_it_from_unapplied(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            plus_extra = root / "plus-extra.xlsx"
            label = "Plus Ek İndirim %10"
            write_workbook(
                plus_extra,
                ["Barkod", "Maksimum Girebileceğin Fiyat", "Kampanyalı Satış Fiyatı"],
                [["E-1", 100, None]],
            )
            (root / "manifest.json").write_text(
                json.dumps({
                    "files": {},
                    "plus_extra_configs": [
                        {"path": str(plus_extra), "label": label, "rate": 10}
                    ],
                }),
                encoding="utf-8",
            )
            row = {
                "Barkod": "E-1",
                "Stok Adedi": 1,
                "İlk Kampanya Seçimi": "Hiçbiri",
                "İlk Ekstra Kampanya Seçimi": "Hiçbiri",
                "Uygulanabilir Kampanyalar": "Plus Ek İndirim",
                "eligible_main_campaigns": ["Hiçbiri"],
                "eligible_extra_campaigns": ["Hiçbiri", label],
                "Güncel Ürün Fiyatı (TL)": 100,
                "Güncel Ürün Kalan Net (TL)": 90,
                "Güncel Ürün Komisyon (%)": 10,
                "Plus Ek Fiyatı %10 (TL)": 90,
                "Plus Ek Net %10 (TL)": 81,
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {},
                {
                    "target_type": "Hepsi",
                    "selections": {"E-1": {"main": "Hiçbiri", "extra": label}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            run_dir = output_dir / response.get_json()["timestamp_folder"]
            extra_export = run_dir / "Trendyol_Plus_Musterilerine_Ozel_Ek_%10_Indirim.xlsx"
            self.assertTrue(extra_export.exists())
            general = pd.read_excel(run_dir / "Kampanya_Genel_Raporu.xlsx")
            self.assertEqual(general.loc[0, "Uygulanan Kampanya Seçimi"], "Hiçbiri")
            self.assertEqual(general.loc[0, "Uygulanan Ekstra Kampanya Seçimi"], label)
            unapplied = pd.read_excel(run_dir / "Uygulanmayan_Urunler_Raporu.xlsx")
            self.assertTrue(unapplied.empty)

    def test_plus_extra_only_target_exports_evaluated_customer_price_and_config_name(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            plus_extra = root / "plus-extra.xlsx"
            label = (
                "Plus Ek İndirim (300 TL Üzeri / 50 TL İndirim / "
                "%60 Trendyol Karşılamalı)"
            )
            write_workbook(
                plus_extra,
                ["Barkod", "Maksimum Girebileceğin Fiyat", "Kampanyalı Satış Fiyatı"],
                [["E-1", 400, None]],
            )
            (root / "manifest.json").write_text(
                json.dumps({
                    "files": {},
                    "plus_extra_configs": [{
                        "path": str(plus_extra),
                        "label": label,
                        "min_price": 300,
                        "discount_amount": 50,
                        "discount_type": "TL",
                        "trendyol_percent": 60,
                    }],
                }),
                encoding="utf-8",
            )
            row = {
                "Barkod": "E-1",
                "Stok Adedi": 1,
                "İlk Kampanya Seçimi": "Hiçbiri",
                "İlk Ekstra Kampanya Seçimi": "Hiçbiri",
                "Uygulanabilir Kampanyalar": "Plus Ek İndirim",
                "eligible_main_campaigns": ["Hiçbiri"],
                "eligible_extra_campaigns": ["Hiçbiri", label],
                "counter_evaluations": {
                    label: {
                        "price": 400,
                        "customer_price": 350,
                        "net": 330,
                        "rate": 10,
                        "seller_disc": 20,
                        "min_price": 300,
                        "disc_type": "TL",
                        "disc_val": 50,
                        "trendyol_percent": 60,
                    }
                },
            }

            response, output_dir = self.apply_from_temp(
                root,
                [row],
                {},
                {
                    "target_type": "Plus Ek İndirim",
                    "selections": {"E-1": {"main": "Hiçbiri", "extra": label}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            generated = response.get_json()["generated_files"]
            extra_name = next(Path(name).name for name in generated if "Trendyol_Plus" in name)
            self.assertIn("300_TL_Uzeri", extra_name)
            self.assertIn("50_TL_Indirim", extra_name)
            self.assertIn("%60_Trendyol_Karsilamali", extra_name)
            exported = openpyxl.load_workbook(
                output_dir / response.get_json()["timestamp_folder"] / extra_name
            ).active
            self.assertEqual(exported["C2"].value, 350)

    def test_percent_coupon_export_filename_keeps_discount_type(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            coupon = root / "coupon.xlsx"
            label = "300 TL Üzerine %25 Kupon"
            write_workbook(
                coupon,
                ["Barkod", "Eklenecek Ürünleri Seçiniz"],
                [["C-1", None]],
            )
            (root / "manifest.json").write_text(
                json.dumps({
                    "files": {},
                    "coupon_configs": [{
                        "path": str(coupon),
                        "label": label,
                        "min_price": 300,
                        "discount_amount": 25,
                        "discount_type": "%",
                        "trendyol_percent": 60,
                    }],
                }),
                encoding="utf-8",
            )
            row = {
                "Barkod": "C-1",
                "Stok Adedi": 1,
                "Uygulanabilir Kampanyalar": label,
                "eligible_main_campaigns": ["Hiçbiri"],
                "eligible_extra_campaigns": ["Hiçbiri", label],
            }

            response, _output_dir = self.apply_from_temp(
                root,
                [row],
                {},
                {
                    "target_type": "Hepsi",
                    "selections": {"C-1": {"main": "Hiçbiri", "extra": label}},
                },
            )

            self.assertEqual(response.status_code, 200, response.get_json())
            coupon_name = next(
                Path(name).name
                for name in response.get_json()["generated_files"]
                if "Kupon" in name
            )
            self.assertIn("_%25_Kupon_", coupon_name)

    def test_selection_payload_and_number_validation_fail_closed(self):
        self.assertFalse(
            app.selection_payload_is_valid({"A1": {"main": "Bilinmeyen", "extra": "Hiçbiri"}})
        )
        self.assertFalse(
            app.selection_payload_is_valid({"A1": {"main": "Avantajlı", "unexpected": "x"}})
        )
        for value in ("nan", float("nan"), "inf", float("inf"), "-inf"):
            with self.subTest(value=value):
                self.assertIsNone(app.as_number(value))

    def test_calculate_rejects_non_finite_campaign_config_with_validation_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            manifest = root / "manifest.json"
            manifest.write_text('{"files": {}}', encoding="utf-8")
            upload = root / "uploads"
            output = root / "output"
            upload.mkdir()
            output.mkdir()
            with (
                patch.object(app, "UPLOAD_DIR", str(upload)),
                patch.object(app, "OUTPUT_DIR", str(output)),
                patch.object(app, "INPUT_MANIFEST", str(manifest)),
                patch.object(app, "save_upload_set", return_value={
                    "discount": "discount.xlsx",
                    "commission": "commission.xlsx",
                    "current": "current.xlsx",
                }),
                patch(
                    "komisyon_hesaplayici.calculate_all",
                    return_value={"success": False, "message": "should not run"},
                ),
            ):
                response = self.client.post(
                    "/api/calculate",
                    data={
                        "plus_extra_configs_json": json.dumps(
                            [{"rate": float("nan")}]
                        )
                    },
                )

        self.assertEqual(response.status_code, 400)
        self.assertIn("sonlu", response.get_json()["message"])

    def test_calculate_rejects_missing_discount_even_if_loader_returns_other_base_inputs(self):
        with patch.object(
            app,
            "save_upload_set",
            return_value={"commission": "commission.xlsx", "current": "current.xlsx"},
        ):
            response = self.client.post("/api/calculate", data={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("İndirim", response.get_json()["message"])


if __name__ == "__main__":
    unittest.main()
