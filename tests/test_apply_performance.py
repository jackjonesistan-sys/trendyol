import unittest
from unittest.mock import patch

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

import app


class ApplyPerformanceTests(unittest.TestCase):
    def test_safe_keep_rows_compacts_then_deletes_tail_once(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Barkod", "Fiyat", "Formul"])
        for row in range(2, 22):
            sheet.append([f"B-{row}", row * 10, f"=B{row}*2"])

        keep_rows = set(range(2, 22, 2))
        original_delete_rows = sheet.delete_rows
        original_max_column = Worksheet.max_column
        max_column_reads = 0

        def tracked_max_column(worksheet):
            nonlocal max_column_reads
            max_column_reads += 1
            return original_max_column.fget(worksheet)

        with (
            patch.object(sheet, "delete_rows", wraps=original_delete_rows) as delete_rows,
            patch.object(Worksheet, "max_column", new=property(tracked_max_column)),
        ):
            app.safe_keep_rows(sheet, keep_rows)

        self.assertLessEqual(delete_rows.call_count, 1)
        self.assertLessEqual(max_column_reads, 4)
        self.assertEqual(sheet.max_row, 11)
        for output_row, source_row in enumerate(range(2, 22, 2), 2):
            self.assertEqual(sheet.cell(output_row, 1).value, f"B-{source_row}")
            self.assertEqual(sheet.cell(output_row, 2).value, source_row * 10)
            self.assertEqual(sheet.cell(output_row, 3).value, f"=B{output_row}*2")


if __name__ == "__main__":
    unittest.main()
