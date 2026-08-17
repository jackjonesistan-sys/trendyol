import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd

import app
from input_files import (
    build_campaign_label,
    load_net_discount_config,
    normalize_campaign_config,
    save_net_discount_config,
)
from komisyon_hesaplayici import calculate_all


def write_test_workbook(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)


class NetDiscountTests(unittest.TestCase):
    def setUp(self):
        app.app.config["TESTING"] = True
        self.client = app.app.test_client()

    def test_net_discount_config_normalization_and_label(self):
        conf_pct = normalize_campaign_config(
            {"type": "%", "amount": "15", "min_cart_amount": "0", "enabled": True},
            "net_discount",
        )
        self.assertEqual(conf_pct["discount_type"], "%")
        self.assertEqual(conf_pct["discount_amount"], 15.0)
        self.assertEqual(conf_pct["trendyol_percent"], 0.0)
        self.assertEqual(build_campaign_label("net_discount", conf_pct), "%15 Net İndirim")

        conf_tl = normalize_campaign_config(
            {"type": "TL", "amount": 50, "min_cart_amount": 100, "enabled": True},
            "net_discount",
        )
        self.assertEqual(conf_tl["discount_type"], "TL")
        self.assertEqual(conf_tl["discount_amount"], 50.0)
        self.assertEqual(build_campaign_label("net_discount", conf_tl), "50 TL Net İndirim")

    def test_net_discount_manifest_persistence(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            manifest_path = Path(temp_dir) / "manifest.json"
            config_in = {
                "discount_type": "%",
                "discount_amount": 10.0,
                "min_cart_amount": 0.0,
                "order_limit": 5,
                "end_date": "2026-12-31 23:59",
                "enabled": True,
            }
            save_net_discount_config(manifest_path, config_in)
            config_out = load_net_discount_config(manifest_path)
            self.assertEqual(config_out["discount_type"], "%")
            self.assertEqual(config_out["discount_amount"], 10.0)
            self.assertEqual(config_out["order_limit"], 5)
            self.assertEqual(config_out["end_date"], "2026-12-31 23:59")
            self.assertTrue(config_out["enabled"])

    def test_net_discount_calculation_evaluates_seller_net_and_customer_price(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            discount_p = root / "discount.xlsx"
            commission_p = root / "commission.xlsx"
            current_p = root / "current.xlsx"
            net_discount_p = root / "net_discount.xlsx"

            write_test_workbook(
                discount_p,
                ["BARKOD", "Eski Fiyat", "YENİ Fiyat"],
                [["BAR1", 100.0, 80.0]],
            )
            write_test_workbook(
                commission_p,
                ["BARKOD", "1.Fiyat Alt Limit", "1.KOMİSYON", "TARİFE GRUBU"],
                [["BAR1", 0, 15.0, "G1"]],
            )
            write_test_workbook(
                current_p,
                ["Barkod", "Komisyon Oranı", "Piyasa Satış Fiyatı (KDV Dahil)", "Trendyol'da Satılacak Fiyat (KDV Dahil)"],
                [["BAR1", 15.0, 100.0, 100.0]],
            )
            write_test_workbook(
                net_discount_p,
                ["Barkod"],
                [["BAR1"]],
            )

            inputs = {
                "discount": str(discount_p),
                "commission": str(commission_p),
                "current": str(current_p),
                "net_discount": str(net_discount_p),
            }

            net_conf = {
                "discount_type": "%",
                "discount_amount": 10.0,
                "min_cart_amount": 0.0,
                "order_limit": 10,
                "end_date": "",
                "enabled": True,
                "path": str(net_discount_p),
            }

            calc_result = calculate_all(
                input_files=inputs,
                net_discount_config=net_conf,
                output_dir=root / "output",
            )

            self.assertTrue(calc_result.get("success"), calc_result)
            rows = calc_result.get("results", [])
            self.assertEqual(len(rows), 1)
            row = rows[0]

            evals = row.get("counter_evaluations")
            label = "%10 Net İndirim"
            self.assertIn(label, evals)
            self.assertIn(label, row.get("eligible_extra_campaigns"))

            eval_info = evals[label]
            self.assertEqual(eval_info["trendyol_percent"], 0.0)
            self.assertEqual(eval_info["disc_type"], "%")
            self.assertEqual(eval_info["disc_val"], 10.0)
            # Commission on price (100 * 15% = 15), seller discount = 10 -> net = 100 - 15 - 10 = 75.0
            self.assertEqual(eval_info["net"], 75.0)

    def test_net_discount_apply_generates_template_output(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output_dir = root / "output"
            output_dir.mkdir()
            template_path = root / "net_discount.xlsx"
            write_test_workbook(
                template_path,
                [
                    "Trendyol Ürün ID",
                    "Ürün Bilgisi",
                    "Marka",
                    "Renk",
                    "Barkod",
                    "Model Kodu",
                    "Güncel Satış Fiyatı",
                    "Buybox",
                    "Kampayaya Dahil Edilsin Mi?",
                ],
                [
                    [101, "Ürün 1", "M1", "Siyah", "BAR1", "MOD1", 100, 100, "Hayır"],
                    [102, "Ürün 2", "M1", "Beyaz", "BAR2", "MOD2", 200, 200, "Hayır"],
                ],
            )

            manifest_path = root / "manifest.json"
            manifest_path.write_text(
                '{"files": {"net_discount": {"stored_path": "'
                + str(template_path).replace("\\", "\\\\")
                + '"}}, "net_discount_config": {"discount_type": "%", "discount_amount": 10, "min_cart_amount": 0, "enabled": true, "stored_path": "'
                + str(template_path).replace("\\", "\\\\")
                + '"}}',
                encoding="utf-8",
            )

            row1 = {
                "Barkod": "BAR1",
                "Stok Adedi": 1,
                "eligible_main_campaigns": ["Hiçbiri"],
                "eligible_extra_campaigns": ["Hiçbiri", "%10 Net İndirim"],
                "campaign_floor_prices": {},
                "counter_evaluations": {
                    "%10 Net İndirim": {
                        "price": 100,
                        "customer_price": 90,
                        "net": 76.5,
                        "disc_type": "%",
                        "disc_val": 10,
                        "trendyol_percent": 0,
                    }
                },
                "flash_evaluations": [],
            }
            row2 = {
                "Barkod": "BAR2",
                "Stok Adedi": 1,
                "eligible_main_campaigns": ["Hiçbiri"],
                "eligible_extra_campaigns": ["Hiçbiri", "%10 Net İndirim"],
                "campaign_floor_prices": {},
                "counter_evaluations": {
                    "%10 Net İndirim": {
                        "price": 200,
                        "customer_price": 180,
                        "net": 153,
                        "disc_type": "%",
                        "disc_val": 10,
                        "trendyol_percent": 0,
                    }
                },
                "flash_evaluations": [],
            }

            result_path = root / "result.xlsx"
            pd.DataFrame([row1, row2]).to_excel(result_path, index=False)

            with (
                patch.object(app, "OUTPUT_DIR", str(output_dir)),
                patch.object(app, "F_HESAP", str(result_path)),
                patch.object(app, "INPUT_MANIFEST", str(manifest_path)),
                patch.object(app, "load_upload_set", return_value={"net_discount": str(template_path)}),
                patch.object(app, "fix_xlsx_for_trendyol", side_effect=lambda _p: None),
                patch("fiyat_farki_analiz_script.generate_fiyat_farki_raporu", side_effect=lambda *_a, **_k: None),
            ):
                # Select BAR1 for Net Discount, BAR2 for Hiçbiri
                response = self.client.post(
                    "/api/apply",
                    json={
                        "target_type": "Net İndirim",
                        "selections": {
                            "BAR1": {"main": "Hiçbiri", "extra": "%10 Net İndirim"},
                            "BAR2": {"main": "Hiçbiri", "extra": "Hiçbiri"},
                        },
                    },
                )

            self.assertEqual(response.status_code, 200, response.get_json())
            data = response.get_json()
            self.assertTrue(data["success"])

            # Check generated files in the output run folder
            created_files = list(output_dir.glob("*/*.xlsx"))
            net_discount_files = [f for f in created_files if "Net_Indirim" in f.name]
            self.assertEqual(len(net_discount_files), 1)

            out_wb = openpyxl.load_workbook(net_discount_files[0])
            out_ws = out_wb.active

            # Header row + 1 product row (BAR1)
            self.assertEqual(out_ws.max_row, 2)
            self.assertEqual(out_ws.cell(2, 5).value, "BAR1")  # Barkod column
            self.assertEqual(out_ws.cell(2, 9).value, "Evet")  # Kampayaya Dahil Edilsin Mi? column


if __name__ == "__main__":
    unittest.main()
