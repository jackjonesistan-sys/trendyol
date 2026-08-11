import tempfile
import unittest
from pathlib import Path

import openpyxl

from xlsx_postprocess import fix_xlsx_for_trendyol


class XlsxPostprocessCachedFormulaTests(unittest.TestCase):
    def _process(self, temp_dir, filename, workbook):
        path = Path(temp_dir) / filename
        workbook.save(path)
        fix_xlsx_for_trendyol(str(path))
        return openpyxl.load_workbook(path, data_only=True)

    def test_multi_period_formulas_use_their_referenced_columns_and_tariffs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = {
                "M": "Plus Fiyat Üst Limiti",
                "O": "Plus Komisyon Teklifi",
                "Q": "Plus Komisyon Teklifi",
                "S": "Plus Fiyat Seçimi",
                "T": "Tarife Seçimi",
                "V": "Hesaplanan Komisyon (3 Gün)",
                "W": "Hesaplanan Komisyon (4 Gün)",
                "X": "İptal",
            }
            for column, value in headers.items():
                sheet[f"{column}1"] = value

            rows = (
                (100, 13.1, 14.1, 99, "7 Günlük Fiyat"),
                (100, 23.2, 24.2, 90, "3 Günlük Fiyat (ilk aralık)"),
                (100, 33.3, 34.3, 90, "4 Günlük Fiyat (ikinci aralık)"),
                (100, 43.4, 44.4, 90, None),
                (100, 53.5, 54.5, 101, "7 Günlük Fiyat"),
                (100, 63.6, 64.6, None, "7 Günlük Fiyat"),
                (100, None, 74.7, 90, "7 Günlük Fiyat"),
                (None, 83.8, 84.8, 90, "7 Günlük Fiyat"),
            )
            for row, (upper, first_offer, second_offer, price, tariff) in enumerate(rows, 2):
                sheet[f"M{row}"] = upper
                sheet[f"O{row}"] = first_offer
                sheet[f"Q{row}"] = second_offer
                sheet[f"S{row}"] = price
                sheet[f"T{row}"] = tariff
                sheet[f"V{row}"] = (
                    f'=IF(OR(S{row}="", T{row}="", ISBLANK(O{row})), "-", '
                    f'IF(OR(ISNUMBER(SEARCH("3 Günlük", T{row})), '
                    f'ISNUMBER(SEARCH("7 Günlük", T{row}))), '
                    f'IF(S{row}<=M{row}, O{row}, "-"), "-"))'
                )
                sheet[f"W{row}"] = (
                    f'=IF(OR(S{row}="", T{row}=""), "-", '
                    f'IF(OR(ISNUMBER(SEARCH("4 Günlük", T{row})), '
                    f'ISNUMBER(SEARCH("7 Günlük", T{row}))), '
                    f'IF(S{row}<=M{row}, Q{row}, "-"), "-"))'
                )
                sheet[f"X{row}"] = f'=IF(ISBLANK(S{row}), "", "Hayır")'

            processed = self._process(temp_dir, "multi-period.xlsx", workbook)
            try:
                values = processed.active
                self.assertEqual(
                    [
                        (
                            values[f"V{row}"].value,
                            values[f"W{row}"].value,
                            values[f"X{row}"].value,
                        )
                        for row in range(2, 10)
                    ],
                    [
                        (13.1, 14.1, "Hayır"),
                        (23.2, "-", "Hayır"),
                        ("-", 34.3, "Hayır"),
                        ("-", "-", "Hayır"),
                        ("-", "-", "Hayır"),
                        ("-", "-", None),
                        ("-", 74.7, "Hayır"),
                        ("-", "-", "Hayır"),
                    ],
                )
            finally:
                processed.close()

    def test_legacy_seven_day_cached_values_are_preserved(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["M1"] = "Plus Fiyat Üst Limiti"
            sheet["O1"] = "Plus Komisyon Teklifi"
            sheet["Q1"] = "Plus Fiyat Seçimi"
            sheet["R1"] = "Tarife Seçimi"
            sheet["S1"] = "Hesaplanan Komisyon (7 Gün)"
            sheet["T1"] = "İptal"
            sheet["M2"] = 100
            sheet["O2"] = 17.5
            sheet["Q2"] = 95
            sheet["R2"] = "7 Günlük Fiyat"
            sheet["S2"] = (
                '=IF(OR(Q2="", R2=""), "-", IF(ISNUMBER(SEARCH("7 Günlük", R2)), '
                'IF(Q2<=M2, O2, "-"), "-"))'
            )
            sheet["T2"] = '=IF(ISBLANK(Q2), "", "Hayır")'

            processed = self._process(temp_dir, "legacy.xlsx", workbook)
            try:
                values = processed.active
                self.assertEqual(values["S2"].value, 17.5)
                self.assertEqual(values["T2"].value, "Hayır")
            finally:
                processed.close()


if __name__ == "__main__":
    unittest.main()
