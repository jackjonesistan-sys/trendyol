import os
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
import pandas as pd
import openpyxl

from komisyon_hesaplayici import (
    calculate_all,
    evaluate_commission_tariff_bracket,
    find_commission_tariff_period_label,
)
import app


class CommissionTariffIntegrationTests(unittest.TestCase):
    def setUp(self):
        app.app.config["TESTING"] = True
        self.client = app.app.test_client()

    def test_evaluate_commission_tariff_bracket_selects_lowest_eligible_customer_price(self):
        # 4 brackets:
        # 4. Kademe: 800 TL @ 10% -> Net: 720 TL
        # 3. Kademe: 850 TL @ 12% -> Net: 748 TL
        # 2. Kademe: 900 TL @ 15% -> Net: 765 TL
        # 1. Kademe: 950 TL @ 18% -> Net: 779 TL
        kom_row = {
            "Barkod": "TEST-1",
            "1.Fiyat Alt Limit": 950,
            "1.KOMİSYON": 18,
            "2.Fiyat Üst Limiti": 900,
            "2.KOMİSYON": 15,
            "3.Fiyat Üst Limiti": 850,
            "3.KOMİSYON": 12,
            "4.Fiyat Üst Limiti": 800,
            "4.KOMİSYON": 10,
            "Tarife Seçimi": "7 Günlük Fiyat",
        }

        # Case 1: Dip Price = 750, Dip Net = 650 -> All 4 brackets eligible.
        # Should pick lowest customer price = 4. Kademe (800 TL).
        eval_1 = evaluate_commission_tariff_bracket(kom_row, dip_price=750, dip_net=650)
        self.assertIsNotNone(eval_1)
        self.assertTrue(eval_1["has_eligible"])
        self.assertEqual(eval_1["price"], 800)
        self.assertEqual(eval_1["rate"], 10)
        self.assertEqual(eval_1["net"], 720)
        self.assertEqual(eval_1["kademe_no"], 4)
        self.assertEqual(eval_1["tariff_selection"], "7 Günlük Fiyat")

        # Case 2: Dip Price = 820, Dip Net = 700 -> 4. Kademe price (800 TL) is below Dip Price (820 TL),
        # but its net (720 TL) is HIGHER than Dip Net (700 TL).
        # It must be eligible and chosen because it gives the customer the lowest price while preserving profitability!
        eval_2 = evaluate_commission_tariff_bracket(kom_row, dip_price=820, dip_net=700)
        self.assertIsNotNone(eval_2)
        self.assertTrue(eval_2["has_eligible"])
        self.assertEqual(eval_2["price"], 800)
        self.assertEqual(eval_2["kademe_no"], 4)
        self.assertEqual(eval_2["rate"], 10)
        self.assertEqual(eval_2["net"], 720)

        # Case 3: Dip Price = 750, Dip Net = 760 -> Kademe 4 (Net 720) and Kademe 3 (Net 748) fail dip net.
        # Eligible brackets: 2 (Net 765), 1 (Net 779). Lowest price = 2. Kademe (900 TL).
        eval_3 = evaluate_commission_tariff_bracket(kom_row, dip_price=750, dip_net=760)
        self.assertIsNotNone(eval_3)
        self.assertTrue(eval_3["has_eligible"])
        self.assertEqual(eval_3["price"], 900)
        self.assertEqual(eval_3["kademe_no"], 2)

    def test_find_commission_tariff_period_label_extraction(self):
        row_with_explicit = {"Tarife Seçimi": "14 Günlük Fiyat"}
        self.assertEqual(find_commission_tariff_period_label(row_with_explicit), "14 Günlük Fiyat")

        row_with_col = {"Tarih aralığı (15 Gün)": "01.08-15.08"}
        self.assertEqual(find_commission_tariff_period_label(row_with_col), "15 Günlük Fiyat")

        row_default = {"Barkod": "123"}
        self.assertEqual(find_commission_tariff_period_label(row_default), "7 Günlük Fiyat")

    def test_calculate_all_and_export_commission_tariff(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            discount = root / "discount.xlsx"
            commission = root / "commission.xlsx"
            current = root / "current.xlsx"
            output = root / "output"

            pd.DataFrame([
                {"BARKOD": "BARKOD-1", "Eski Fiyat": 1000, "YENİ Fiyat": 700, "Durum": "İndirim"},
                {"BARKOD": "BARKOD-2", "Eski Fiyat": 1000, "YENİ Fiyat": 700, "Durum": "İndirim"},
            ]).to_excel(discount, index=False)

            pd.DataFrame([
                {
                    "Barkod": "BARKOD-1",
                    "1.Fiyat Alt Limit": 950,
                    "1.KOMİSYON": 20,
                    "2.Fiyat Üst Limiti": 900,
                    "2.KOMİSYON": 15,
                    "3.Fiyat Üst Limiti": 850,
                    "3.KOMİSYON": 12,
                    "4.Fiyat Üst Limiti": 800,
                    "4.KOMİSYON": 10,
                    "Tarih aralığı (7 Gün)": "17.08-24.08",
                    "YENİ TSF (FİYAT GÜNCELLE)": None,
                    "Tarife Seçimi": None,
                },
                {
                    "Barkod": "BARKOD-2",
                    "1.Fiyat Alt Limit": 950,
                    "1.KOMİSYON": 20,
                    "2.Fiyat Üst Limiti": 900,
                    "2.KOMİSYON": 15,
                    "3.Fiyat Üst Limiti": 850,
                    "3.KOMİSYON": 12,
                    "4.Fiyat Üst Limiti": 800,
                    "4.KOMİSYON": 10,
                    "Tarih aralığı (7 Gün)": "17.08-24.08",
                    "YENİ TSF (FİYAT GÜNCELLE)": None,
                    "Tarife Seçimi": None,
                },
            ]).to_excel(commission, index=False)

            pd.DataFrame([
                {
                    "Barkod": "BARKOD-1",
                    "Komisyon Oranı": 20,
                    "Piyasa Satış Fiyatı (KDV Dahil)": 1200,
                    "Trendyol'da Satılacak Fiyat (KDV Dahil)": 1000,
                },
                {
                    "Barkod": "BARKOD-2",
                    "Komisyon Oranı": 20,
                    "Piyasa Satış Fiyatı (KDV Dahil)": 1200,
                    "Trendyol'da Satılacak Fiyat (KDV Dahil)": 1000,
                },
            ]).to_excel(current, index=False)

            result = calculate_all(
                {
                    "discount": discount,
                    "commission": commission,
                    "current": current,
                },
                output_dir=output,
            )

            self.assertTrue(result["success"])
            rows = {r["Barkod"]: r for r in result["results"]}

            # Check BARKOD-1 calculation fields
            row1 = rows["BARKOD-1"]
            self.assertEqual(row1["Komisyon Tarifesi Fiyatı (TL)"], 800)
            self.assertEqual(row1["Komisyon Tarifesi Komisyon (%)"], 10)
            self.assertEqual(row1["Komisyon Tarifesi Net (TL)"], 720)
            self.assertEqual(row1["Komisyon Tarifesi Seçimi"], "7 Günlük Fiyat")
            self.assertIn("Komisyon Tarifesi", row1["eligible_campaigns"])
            self.assertIn("Komisyon Tarifesi", row1["all_matching_main_campaigns"])

            # Test Exporting Urun_Komisyon_Tarifeleri_Urunler.xlsx via apply API
            # User selects Komisyon Tarifesi for BARKOD-1, and Hiçbiri for BARKOD-2
            manifest_path = root / "manifest.json"
            manifest_path.write_text(
                json.dumps({
                    "files": {
                        "commission": {
                            "stored_name": commission.name,
                            "original_name": "Ürün Komisyon Tarifeleri.xlsx",
                        }
                    },
                    "user_selections": {},
                }),
                encoding="utf-8",
            )

            result_path = output / "Kampanya_Hesaplama_Sonuclari.xlsx"
            with (
                patch.object(app, "UPLOAD_DIR", str(root)),
                patch.object(app, "OUTPUT_DIR", str(output)),
                patch.object(app, "F_HESAP", str(result_path)),
                patch.object(app, "INPUT_MANIFEST", str(manifest_path)),
            ):
                resp = self.client.post(
                    "/api/apply",
                    json={
                        "target_type": "Komisyon Tarifesi",
                        "selections": {
                            "BARKOD-1": {"main": "Komisyon Tarifesi", "extra": "Hiçbiri"},
                            "BARKOD-2": {"main": "Hiçbiri", "extra": "Hiçbiri"},
                        },
                        "ignore_zero_stock": True,
                    },
                )
                self.assertEqual(resp.status_code, 200)
                data = resp.get_json()
                self.assertTrue(data.get("success"), data)

                out_files = data.get("generated_files", [])
                self.assertTrue(any("Urun_Komisyon_Tarifeleri_Urunler.xlsx" in f for f in out_files))

                # Verify exported file contents
                export_file = output / out_files[0]
                wb = openpyxl.load_workbook(export_file)
                ws = wb.active
                header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                self.assertIn("Barkod", header)
                self.assertIn("YENİ TSF (FİYAT GÜNCELLE)", header)
                self.assertIn("Tarife Seçimi", header)

                b_idx = header.index("Barkod") + 1
                p_idx = header.index("YENİ TSF (FİYAT GÜNCELLE)") + 1
                t_idx = header.index("Tarife Seçimi") + 1

                # Only BARKOD-1 should remain in the exported file
                barcodes_in_export = [ws.cell(r, b_idx).value for r in range(2, ws.max_row + 1)]
                self.assertEqual(barcodes_in_export, ["BARKOD-1"])
                self.assertEqual(ws.cell(2, p_idx).value, 800.0)
                self.assertEqual(ws.cell(2, t_idx).value, "7 Günlük Fiyat")

    def test_advantage_campaign_inherits_more_profitable_commission_tariff_price(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            discount = root / "discount.xlsx"
            commission = root / "commission.xlsx"
            current = root / "current.xlsx"
            advantage = root / "advantage.xlsx"
            output = root / "output"

            # Dip limit = 3389.55
            pd.DataFrame([
                {"BARKOD": "BARKOD-WIN", "Eski Fiyat": 4904, "YENİ Fiyat": 3389.55, "Durum": "İndirim"},
            ]).to_excel(discount, index=False)

            # Commission tariff has a 3rd bracket at 3380.27 (cheaper price, but commission drops to 20.4% yielding higher net: 2690.69 vs 2667.58)
            pd.DataFrame([
                {
                    "Barkod": "BARKOD-WIN",
                    "1.Fiyat Alt Limit": 3602.40,
                    "1.KOMİSYON": 22.0,
                    "2.Fiyat Üst Limiti": 3602.39,
                    "2.KOMİSYON": 21.3,
                    "3.Fiyat Üst Limiti": 3380.27,
                    "3.KOMİSYON": 20.4,
                    "4.Fiyat Üst Limiti": 3079.98,
                    "4.KOMİSYON": 19.0,
                    "Tarih aralığı (7 Gün)": "18.08-25.08",
                },
            ]).to_excel(commission, index=False)

            pd.DataFrame([
                {
                    "Barkod": "BARKOD-WIN",
                    "Komisyon Oranı": 22.0,
                    "Piyasa Satış Fiyatı (KDV Dahil)": 6000,
                    "Trendyol'da Satılacak Fiyat (KDV Dahil)": 4904,
                },
            ]).to_excel(current, index=False)

            # Advantage list has 3389.55
            pd.DataFrame([
                {
                    "BARKOD": "BARKOD-WIN",
                    "YENİ TSF (FİYAT GÜNCELLE)": 3389.55,
                    "1 YILDIZ ÜST FİYAT": 3539.55,
                },
            ]).to_excel(advantage, index=False)

            result = calculate_all(
                {
                    "discount": discount,
                    "commission": commission,
                    "current": current,
                    "advantage": advantage,
                },
                output_dir=output,
            )

            self.assertTrue(result["success"])
            row = result["results"][0]

            # Avantajlı must inherit the superior tariff price and rate
            self.assertEqual(row["Avantajlı Ürün Fiyatı (YENİ TSF) (TL)"], 3380.27)
            self.assertEqual(row["Avantajlı Ürün Komisyon (%)"], 20.4)
            self.assertEqual(row["Avantajlı Ürün Kalan Net (TL)"], 2690.69)
            self.assertEqual(row["Komisyon Tarifesi Fiyatı (TL)"], 3380.27)
            self.assertEqual(row["Komisyon Tarifesi Net (TL)"], 2690.69)
            self.assertIn("Avantajlı", row["eligible_campaigns"])


if __name__ == "__main__":
    unittest.main()
