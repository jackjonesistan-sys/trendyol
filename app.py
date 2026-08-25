import math

def clean_nans(obj):
    """Recursively replaces NaN, Infinity, -Infinity values with None for standard JSON serialization."""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    elif isinstance(obj, dict):
        return {k: clean_nans(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nans(v) for v in obj]
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    return obj

import ast
import os
import json
import math
import re
from io import BytesIO
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request, send_from_directory
from input_files import (
    INPUT_SPECS,
    InputValidationError,
    load_upload_set,
    load_upload_status,
    save_upload_set,
    load_counter_configs,
    save_counter_configs,
    load_plus_extra_configs,
    save_plus_extra_configs,
    load_coupon_configs,
    save_coupon_configs,
    load_net_discount_config,
    save_net_discount_config,
    save_single_file_enabled,
    load_recommendation_rule,
    save_single_file_expiries,
    load_user_selections,
    build_campaign_label,
    find_plus_period_columns,
    normalize_campaign_config,
    normalize_campaign_configs,
    normalize_recommendation_rule,
    save_recommendation_rule,
    save_user_selections,
)
from xlsx_postprocess import fix_xlsx_for_trendyol

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 240 * 1024 * 1024

try:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
except Exception:
    pass

@app.after_request
def add_custom_headers(response):
    response.headers["ngrok-skip-browser-warning"] = "1"
    return response

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "Girdiler")
OUTPUT_DIR = os.path.join(BASE_DIR, "Çıktılar")
UPLOAD_DIR = os.path.join(INPUT_DIR, "Yuklenen")
INPUT_MANIFEST = os.path.join(INPUT_DIR, "yuklenen_girdiler.json")

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

F_HESAP = os.path.join(OUTPUT_DIR, "Kampanya_Hesaplama_Sonuclari.xlsx")

REPORT_COLUMNS = [
    "Barkod",
    "Güncel Fiyat (TL)",
    "Güncel Net",
    "Güncel Komisyon",
    "Avantajlı Fiyat (TL)",
    "Avantajlı Net",
    "Flaş Fiyat (TL)",
    "Flaş Net",
    "Plus Fiyat (TL)",
    "Plus Net",
    "Plus Ek İndirim Fiyat (TL)",
    "Plus Ek İndirim Net",
    "Karşılamalı Kampanya Fiyat (TL)",
    "Karşılamalı Kampanya Net",
    "Uygulanan Kampanya",
    "Ekstra Kampanya",
    "Hangisi Karlı?",
    "Düşülebilecek Dip Fiyat (TL)",
    "Uygulanan Kampanya Fiyat",
    "Uygulanan Kampanya Net",
    "Uygulanan Kampanya Komisyon",
    "Uygulanabilecek İndirim (TL)",
    "Uygulanabilecek İndirim (%)",
    "Uygulanan İndirim (TL)",
    "Uygulanan İndirim (%)",
    "Ekstra Uygulanabilir İndirim (TL)",
    "Ekstra Uygulanabilir İndirim (%)",
]
ROUNDING_EPSILON = 1e-9

REQUIRED_RESULT_COLUMNS = {
    "Barkod",
    "Güncel Ürün Fiyatı (TL)",
    "Güncel Ürün Komisyon (%)",
    "Güncel Ürün Kalan Net (TL)",
    "Avantajlı Ürün Fiyatı (YENİ TSF) (TL)",
    "Avantajlı Ürün Komisyon (%)",
    "Avantajlı Ürün Kalan Net (TL)",
    "Flaş Ürün 24 Saat Fiyatı (TL)",
    "Flaş Ürün Komisyon (%)",
    "Flaş Ürün Kalan Net (TL)",
    "Plus Fiyatı (TL)",
    "Plus Komisyon (%)",
    "Plus Net (TL)",
    "Komisyon Tarifesi Fiyatı (TL)",
    "Komisyon Tarifesi Komisyon (%)",
    "Komisyon Tarifesi Net (TL)",
    "Uygulanabilir Kampanyalar",
    "Önerilen Kampanya",
    "Önerilen Ekstra Kampanya",
    "İlk Kampanya Seçimi",
    "İlk Ekstra Kampanya Seçimi",
    "Düşülebilecek Dip Fiyat (TL)",
    "eligible_main_campaigns",
    "all_matching_main_campaigns",
    "eligible_extra_campaigns",
    "all_matching_extra_campaigns",
    "eligible_campaigns",
    "all_matching_campaigns",
    "counter_evaluations",
    "flash_evaluations",
    "dip_details",
    "campaign_floor_prices",
}

CAMPAIGN_LABELS = {
    "Avantajlı": "Avantajlı Ürün",
    "Flaş": "Flaş Ürün",
    "Plus": "Plus Ürün",
    "Komisyon Tarifesi": "Komisyon Tarifesi",
}
VALID_TARGET_TYPES = {
    "Hepsi",
    "Avantajlı",
    "Flaş",
    "Plus",
    "Komisyon Tarifesi",
    "Plus Ek İndirim",
    "Karşılamalı Kampanya",
    "Net İndirim",
}
MAIN_SELECTIONS = {"Hiçbiri", "Avantajlı", "Flaş", "Plus", "Komisyon Tarifesi"}


def as_number(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def normalize_flash_period(value):
    normalized = re.sub(r"\s+", " ", str(value or "").strip()).casefold()
    if normalized in {"24 saat", "24 saat fiyat"}:
        return "24 Saat"
    if normalized in {"3 saat", "3 saat fiyat"}:
        return "3 Saat"
    return None


def normalize_flash_interval_value(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")

    text = str(value).strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).replace(microsecond=0).isoformat(sep=" ")
    except ValueError:
        pass
    for date_format in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format).isoformat(sep=" ")
        except ValueError:
            continue
    return re.sub(r"\s+", " ", text).casefold()


def parse_campaign_configs_json(raw_value, campaign_type):
    if not raw_value:
        return []
    try:
        configs = json.loads(raw_value)
    except (TypeError, ValueError) as exc:
        raise InputValidationError("Kampanya yapılandırması okunamadı.") from exc
    return normalize_campaign_configs(configs, campaign_type)


def parse_recommendation_rule_json(raw_value):
    try:
        rule = json.loads(raw_value)
    except (TypeError, ValueError) as exc:
        raise InputValidationError("Öneri kuralı okunamadı.") from exc
    if rule is None:
        raise InputValidationError("Öneri kuralı boş olamaz.")
    return normalize_recommendation_rule(rule)


def parse_persisted_collection(value, expected_type):
    if isinstance(value, expected_type):
        return value
    if isinstance(value, str):
        try:
            value = ast.literal_eval(value)
        except (SyntaxError, ValueError):
            pass
    return value if isinstance(value, expected_type) else expected_type()


def restore_persisted_collections(frame):
    return frame.assign(**{
        column: frame[column].map(
            lambda value, kind=expected_type: parse_persisted_collection(value, kind)
        )
        for column, expected_type in (
            ("eligible_main_campaigns", list),
            ("all_matching_main_campaigns", list),
            ("eligible_extra_campaigns", list),
            ("all_matching_extra_campaigns", list),
            ("eligible_campaigns", list),
            ("all_matching_campaigns", list),
            ("counter_evaluations", dict),
            ("flash_evaluations", list),
            ("dip_details", list),
            ("campaign_floor_prices", dict),
        )
        if column in frame.columns
    })


def normalize_selection(selection):
    if isinstance(selection, str):
        value = selection.strip()
        if not value:
            return None
        return (value, "Hiçbiri") if value in MAIN_SELECTIONS else ("Hiçbiri", value)
    if not isinstance(selection, dict) or set(selection) - {"main", "extra"}:
        return None
    main = selection.get("main", "Hiçbiri")
    extra = selection.get("extra", "Hiçbiri")
    if (
        not isinstance(main, str)
        or not isinstance(extra, str)
        or main not in MAIN_SELECTIONS
        or not extra.strip()
    ):
        return None
    return main, extra.strip()


def selection_payload_is_valid(selections):
    return isinstance(selections, dict) and all(
        isinstance(barcode, str)
        and bool(barcode.strip())
        and normalize_selection(selection) is not None
        for barcode, selection in selections.items()
    )


def round2(value):
    return math.floor((float(value) * 100) + 0.5 + ROUNDING_EPSILON) / 100


def discounted_price(value, rate):
    price = as_number(value)
    return round2(price * (1 - rate / 100)) if price is not None else None


def plus_extra_export_filename(config, index=0):
    config = normalize_campaign_config(config, "plus_extra")
    format_number = lambda value: (
        str(int(value)) if float(value).is_integer() else str(float(value))
    )
    discount = format_number(config["discount_amount"])
    if (
        config["discount_type"] == "%"
        and config["min_price"] == 0
        and config["trendyol_percent"] == 0
    ):
        return f"Trendyol_Plus_Musterilerine_Ozel_Ek_%{discount}_Indirim.xlsx"

    min_price = format_number(config["min_price"])
    trendyol = format_number(config["trendyol_percent"])
    discount_part = (
        f"%{discount}" if config["discount_type"] == "%" else f"{discount}_TL"
    )
    return (
        "Trendyol_Plus_Musterilerine_Ozel_Ek_"
        f"{min_price}_TL_Uzeri_{discount_part}_Indirim_"
        f"%{trendyol}_Trendyol_Karsilamali.xlsx"
    )


def discount_between(current_price, campaign_price):
    current = as_number(current_price)
    campaign = as_number(campaign_price)
    if current is None or campaign is None or current <= 0 or campaign <= 0:
        return None, None
    amount = round2(max(current - campaign, 0))
    return amount, round2((amount / current) * 100)


def selected_campaign_values(row):
    main_sel = row.get("userSelection", "Hiçbiri") or "Hiçbiri"
    extra_sel = row.get("userExtraSelection", "Hiçbiri") or "Hiçbiri"

    if main_sel == "Avantajlı":
        base_price = as_number(row.get("Avantajlı Ürün Fiyatı (YENİ TSF) (TL)"))
        base_net = as_number(row.get("Avantajlı Ürün Kalan Net (TL)"))
        base_comm = as_number(row.get("Avantajlı Ürün Komisyon (%)"))
    elif main_sel == "Flaş":
        base_price = as_number(row.get("Flaş Ürün 24 Saat Fiyatı (TL)"))
        base_net = as_number(row.get("Flaş Ürün Kalan Net (TL)"))
        base_comm = as_number(row.get("Flaş Ürün Komisyon (%)"))
    elif main_sel == "Plus":
        base_price = as_number(row.get("Plus Fiyatı (TL)"))
        base_net = as_number(row.get("Plus Net (TL)"))
        base_comm = as_number(row.get("Plus Komisyon (%)"))
    elif main_sel == "Komisyon Tarifesi":
        base_price = as_number(row.get("Komisyon Tarifesi Fiyatı (TL)"))
        base_net = as_number(row.get("Komisyon Tarifesi Net (TL)"))
        base_comm = as_number(row.get("Komisyon Tarifesi Komisyon (%)"))
    else:
        base_price = as_number(row.get("Güncel Ürün Fiyatı (TL)"))
        base_net = as_number(row.get("Güncel Ürün Kalan Net (TL)"))
        base_comm = as_number(row.get("Güncel Ürün Komisyon (%)"))

    if base_price is None:
        base_price = as_number(row.get("Güncel Ürün Fiyatı (TL)"))
    if base_comm is None:
        base_comm = as_number(row.get("Güncel Ürün Komisyon (%)"))

    if extra_sel == "Hiçbiri" or not extra_sel:
        return base_price, base_net, base_comm

    counter_evals = row.get("counter_evaluations", {})
    if isinstance(counter_evals, dict) and extra_sel in counter_evals:
        c_info = counter_evals[extra_sel]
        c_price = as_number(c_info.get("price"))
        if c_price is None:
            c_price = base_price
        disc_type = c_info.get("disc_type", "%")
        disc_val = as_number(c_info.get("disc_val", 0))
        trendyol_percent = as_number(c_info.get("trendyol_percent", 0))
        disc_val = 0 if disc_val is None else disc_val
        trendyol_percent = 0 if trendyol_percent is None else trendyol_percent

        if (
            main_sel != "Hiçbiri"
            and base_price is not None
            and c_price is not None
            and base_price < c_price
        ):
            if disc_type == "%":
                tot_disc = round2(base_price * (disc_val / 100.0))
            else:
                tot_disc = disc_val
            seller_disc = round2(tot_disc * (1.0 - (trendyol_percent / 100.0)))
            final_price = round2(base_price - tot_disc)
            final_comm = 0 if base_comm is None else base_comm
            final_net = round2(
                base_price - (base_price * (final_comm / 100.0)) - seller_disc
            )
            return final_price, final_net, final_comm

        final_price = as_number(c_info.get("customer_price"))
        if final_price is None and c_price is not None:
            total_discount = (
                round2(c_price * (disc_val / 100.0))
                if disc_type == "%"
                else disc_val
            )
            final_price = round2(c_price - total_discount)
        final_comm = as_number(c_info.get("rate"))
        if final_comm is None:
            final_comm = 0 if base_comm is None else base_comm
        final_net = as_number(c_info.get("net"))
        if final_net is None and c_price is not None:
            seller_disc = as_number(c_info.get("seller_disc", 0)) or 0
            final_net = round2(
                c_price - (c_price * (final_comm / 100.0)) - seller_disc
            )
        return final_price, final_net, final_comm

    if extra_sel.startswith("Plus Ek İndirim %"):
        try:
            rate = float(extra_sel.rsplit("%", 1)[-1])
            rate_key = str(int(rate)) if rate.is_integer() else str(rate)
            stored_price = as_number(row.get(f"Plus Ek Fiyatı %{rate_key} (TL)"))
            stored_net = as_number(row.get(f"Plus Ek Net %{rate_key} (TL)"))
            final_comm = as_number(row.get("Plus Ek Komisyon (%)"))
            if final_comm is None:
                final_comm = 0 if base_comm is None else base_comm
            if stored_price is not None and stored_net is not None:
                return stored_price, stored_net, final_comm
            total_discount = round2(base_price * (rate / 100.0))
            final_price = round2(base_price - total_discount)
            final_net = round2(
                base_price - (base_price * (final_comm / 100.0)) - total_discount
            )
            return final_price, final_net, final_comm
        except Exception:
            pass

    if extra_sel.endswith("Net İndirim"):
        try:
            if extra_sel.startswith("%"):
                rate = float(extra_sel.replace("%", "").replace("Net İndirim", "").strip())
                disc_type = "%"
            else:
                rate = float(extra_sel.replace("TL", "").replace("Net İndirim", "").strip())
                disc_type = "TL"
            final_comm = 0 if base_comm is None else base_comm
            total_discount = round2(base_price * (rate / 100.0)) if disc_type == "%" else rate
            final_price = round2(base_price - total_discount)
            final_net = round2(
                base_price - (base_price * (final_comm / 100.0)) - total_discount
            )
            return final_price, final_net, final_comm
        except Exception:
            pass

    return base_price, base_net, base_comm


def build_report_row(row):
    main_sel = row.get("userSelection", "Hiçbiri") or "Hiçbiri"
    extra_sel = row.get("userExtraSelection", "Hiçbiri") or "Hiçbiri"
    current_price = as_number(row.get("Güncel Ürün Fiyatı (TL)"))
    campaign_price, campaign_net, campaign_commission = selected_campaign_values(row)
    applied_amount, applied_percent = discount_between(current_price, campaign_price)
    
    dip_price = as_number(row.get("Düşülebilecek Dip Fiyat (TL)"))
    available_amount, available_percent = discount_between(current_price, dip_price)

    extra_amount = None
    extra_percent = None
    if available_amount is not None and applied_amount is not None:
        extra_amount = round2(max(available_amount - applied_amount, 0))
        extra_percent = round2((extra_amount / current_price) * 100) if current_price else None

    plus_extra_price = None
    plus_extra_net = None
    counter_evals = row.get("counter_evaluations", {})
    selected_evaluation = (
        counter_evals.get(extra_sel)
        if isinstance(counter_evals, dict)
        else None
    )
    if extra_sel.startswith("Plus Ek İndirim") and isinstance(selected_evaluation, dict):
        plus_extra_price = campaign_price
        plus_extra_net = campaign_net
    elif extra_sel.startswith("Plus Ek İndirim %"):
        try:
            plus_rate = int(extra_sel.rsplit("%", 1)[-1])
            plus_extra_price = as_number(row.get(f"Plus Ek Fiyatı %{plus_rate} (TL)"))
            plus_extra_net = as_number(row.get(f"Plus Ek Net %{plus_rate} (TL)"))
        except Exception: pass

    counter_price = None
    counter_net = None
    if isinstance(counter_evals, dict) and extra_sel in counter_evals:
        counter_price = campaign_price
        counter_net = campaign_net

    return {
        "Barkod": row.get("Barkod"),
        "Güncel Fiyat (TL)": current_price,
        "Güncel Net": as_number(row.get("Güncel Ürün Kalan Net (TL)")),
        "Güncel Komisyon": as_number(row.get("Güncel Ürün Komisyon (%)")),
        "Avantajlı Fiyat (TL)": as_number(row.get("Avantajlı Ürün Fiyatı (YENİ TSF) (TL)")),
        "Avantajlı Net": as_number(row.get("Avantajlı Ürün Kalan Net (TL)")),
        "Flaş Fiyat (TL)": as_number(row.get("Flaş Ürün 24 Saat Fiyatı (TL)")),
        "Flaş Net": as_number(row.get("Flaş Ürün Kalan Net (TL)")),
        "Plus Fiyat (TL)": as_number(row.get("Plus Fiyatı (TL)")),
        "Plus Net": as_number(row.get("Plus Net (TL)")),
        "Komisyon Tarifesi Fiyat (TL)": as_number(row.get("Komisyon Tarifesi Fiyatı (TL)")),
        "Komisyon Tarifesi Net": as_number(row.get("Komisyon Tarifesi Net (TL)")),
        "Plus Ek İndirim Fiyat (TL)": plus_extra_price,
        "Plus Ek İndirim Net": plus_extra_net,
        "Karşılamalı Kampanya Fiyat (TL)": counter_price,
        "Karşılamalı Kampanya Net": counter_net,
        "Uygulanan Kampanya": CAMPAIGN_LABELS.get(main_sel, main_sel),
        "Ekstra Kampanya": CAMPAIGN_LABELS.get(extra_sel, extra_sel),
        "Hangisi Karlı?": row.get("Hangisi Daha Karlı?"),
        "Düşülebilecek Dip Fiyat (TL)": dip_price,
        "Uygulanan Kampanya Fiyat": campaign_price,
        "Uygulanan Kampanya Net": campaign_net,
        "Uygulanan Kampanya Komisyon": campaign_commission,
        "Uygulanabilecek İndirim (TL)": available_amount,
        "Uygulanabilecek İndirim (%)": available_percent,
        "Uygulanan İndirim (TL)": applied_amount,
        "Uygulanan İndirim (%)": applied_percent,
        "Ekstra Uygulanabilir İndirim (TL)": extra_amount,
        "Ekstra Uygulanabilir İndirim (%)": extra_percent,
    }


def normalize_visible_columns(requested):
    if requested is None:
        return REPORT_COLUMNS.copy()
    if not isinstance(requested, list):
        return REPORT_COLUMNS.copy()
    requested_set = {column for column in requested if isinstance(column, str)}
    return [column for column in REPORT_COLUMNS if column in requested_set]


def campaign_selection_is_applicable(selection, applicable_value):
    if not selection or selection == "Hiçbiri":
        return True
    if not isinstance(selection, str):
        return False
    if selection.endswith("Net İndirim"):
        return True
    applicable = {
        item.strip()
        for item in str(applicable_value or "").split(",")
        if item.strip()
    }
    if selection in applicable:
        return True
    if selection.startswith("Plus Ek İndirim"):
        campaign = "Plus Ek İndirim"
    elif selection.startswith("Karşılamalı"):
        campaign = "Karşılamalı Kampanya"
    elif "Kupon" in selection:
        campaign = "Kupon"
    elif selection.endswith("Net İndirim"):
        campaign = "Net İndirim"
    else:
        campaign = selection
    return campaign in applicable


def row_selection_is_applicable(row, main_selection, extra_selection):
    main_options = row.get("eligible_main_campaigns")
    extra_options = row.get("eligible_extra_campaigns")
    main_ok = (
        main_selection in main_options
        if isinstance(main_options, list) and main_options
        else campaign_selection_is_applicable(
            main_selection, row.get("Uygulanabilir Kampanyalar")
        )
    )
    extra_ok = (
        (extra_selection in extra_options or (extra_selection and extra_selection.endswith("Net İndirim")))
        if isinstance(extra_options, list) and extra_options
        else campaign_selection_is_applicable(
            extra_selection, row.get("Uygulanabilir Kampanyalar")
        )
    )
    return main_ok and extra_ok


REQUIRED_RESULT_COLUMNS = {
    "campaign_floor_prices",
    "eligible_main_campaigns",
    "eligible_extra_campaigns",
    "counter_evaluations",
    "flash_evaluations",
}


def selection_pair_is_safe(row, main_selection, extra_selection):
    main_sel = main_selection or "Hiçbiri"
    extra_sel = extra_selection or "Hiçbiri"
    if not row_selection_is_applicable(row, main_sel, extra_sel):
        return False
    if main_sel == "Hiçbiri" and extra_sel == "Hiçbiri":
        return True

    counter_evals = row.get("counter_evaluations", {})
    if isinstance(counter_evals, dict) and extra_sel in counter_evals:
        evaluation = counter_evals[extra_sel]
        if isinstance(evaluation, dict):
            minimum_price = as_number(evaluation.get("min_price"))
            source_max_price = as_number(evaluation.get("price"))
            main_price_fields = {
                "Avantajlı": "Avantajlı Ürün Fiyatı (YENİ TSF) (TL)",
                "Flaş": "Flaş Ürün 24 Saat Fiyatı (TL)",
                "Plus": "Plus Fiyatı (TL)",
                "Komisyon Tarifesi": "Komisyon Tarifesi Fiyatı (TL)",
            }
            base_price = (
                as_number(row.get(main_price_fields.get(main_sel)))
                if main_sel != "Hiçbiri"
                else as_number(row.get("Güncel Ürün Fiyatı (TL)"))
            )
            if base_price is not None:
                if minimum_price is not None and round2(base_price) < round2(minimum_price):
                    return False
                if source_max_price is not None and round2(base_price) > round2(source_max_price):
                    return False

    final_price, _, _ = selected_campaign_values(
        {**row, "userSelection": main_sel, "userExtraSelection": extra_sel}
    )
    if final_price is not None:
        floor_prices = row.get("campaign_floor_prices", {})
        floor = (
            as_number(floor_prices.get(main_sel))
            if isinstance(floor_prices, dict)
            else None
        )
        if floor is None:
            floor = as_number(row.get("Düşülebilecek Dip Fiyat (TL)"))
        if floor is not None and round2(final_price) < round2(floor):
            return False
    return True


def calculation_result_is_current():
    try:
        return bool(os.path.isfile(F_HESAP) and os.path.getsize(F_HESAP) > 0)
    except OSError:
        return False


def processing_error(label):
    app.logger.exception("%s işlenirken hata", label)
    return jsonify({
        "success": False,
        "message": f"{label} işlenemedi; girdiyi kontrol edip yeniden deneyin.",
    }), 500


def build_report_dataframe(table_data, requested_columns):
    rows = [build_report_row(row) for row in table_data if row.get("Barkod")]
    columns = normalize_visible_columns(requested_columns)
    return pd.DataFrame(rows, columns=REPORT_COLUMNS)[columns]


def write_report_excel(dataframe, output_path):
    dataframe.to_excel(output_path, index=False)
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(dataframe.columns, 1):
        letter = openpyxl.utils.get_column_letter(index)
        sheet.column_dimensions[letter].width = max(len(str(column)) + 2, 12)
    workbook.save(output_path)


@app.route("/api/download/<folder>/<filename>")
def download_file(folder, filename):
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}", folder):
        return jsonify({"success": False, "message": "Çıktı klasörü geçersiz."}), 404
    folder_dir = os.path.join(OUTPUT_DIR, folder)
    return send_from_directory(folder_dir, filename, as_attachment=True)

def safe_keep_rows(ws, keep_row_indices):
    original_max_row = ws.max_row
    keep_rows = sorted({
        row for row in keep_row_indices
        if isinstance(row, int) and 2 <= row <= original_max_row
    })
    if not keep_rows:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        return

    max_column = ws.max_column
    last_column = openpyxl.utils.get_column_letter(max_column)
    for target_row, source_row in enumerate(keep_rows, 2):
        if target_row != source_row:
            ws.move_range(
                f"A{source_row}:{last_column}{source_row}",
                rows=target_row - source_row,
            )

    first_unused_row = len(keep_rows) + 2
    if first_unused_row <= original_max_row:
        ws.delete_rows(first_unused_row, original_max_row - first_unused_row + 1)

    for r in range(2, ws.max_row + 1):
        for c in range(1, max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if isinstance(val, str) and (val.startswith('=') or val.startswith('==')):
                new_val = re.sub(r'([a-zA-Z]+)\d+\b', r'\g<1>' + str(r), val)
                cell.value = new_val


def header_index(ws, name):
    wanted = str(name).strip().casefold()
    for column in range(1, ws.max_column + 1):
        if str(ws.cell(1, column).value or "").strip().casefold() == wanted:
            return column
    return None


def ensure_header(ws, name):
    existing = header_index(ws, name)
    if existing:
        return existing
    column = ws.max_column + 1
    ws.cell(1, column).value = name
    return column


def load_merged_campaign_workbook(standard_path, accounting_path, barcode_header):
    paths = [path for path in (standard_path, accounting_path) if path and os.path.exists(path)]
    if not paths:
        return None

    workbook = openpyxl.load_workbook(paths[0])
    target = workbook.active
    target_barcode_column = header_index(target, barcode_header)
    if not target_barcode_column:
        return workbook

    existing_barcodes = {
        str(target.cell(row, target_barcode_column).value or "").strip()
        for row in range(2, target.max_row + 1)
    }
    target_columns = {
        str(target.cell(1, column).value or "").strip().casefold(): column
        for column in range(1, target.max_column + 1)
    }
    for source_path in paths[1:]:
        source = openpyxl.load_workbook(source_path, data_only=False).active
        source_barcode_column = header_index(source, barcode_header)
        if not source_barcode_column:
            continue
        source_columns = {
            str(source.cell(1, column).value or "").strip().casefold(): column
            for column in range(1, source.max_column + 1)
        }
        shared_columns = {
            target_column: source_columns[name]
            for name, target_column in target_columns.items()
            if name in source_columns
        }
        for row in range(2, source.max_row + 1):
            barcode = str(source.cell(row, source_barcode_column).value or "").strip()
            if not barcode or barcode in existing_barcodes:
                continue
            target_row = target.max_row + 1
            for target_column, source_column in shared_columns.items():
                target.cell(target_row, target_column).value = source.cell(
                    row, source_column
                ).value
            target.cell(target_row, target_barcode_column).value = barcode
            existing_barcodes.add(barcode)
    return workbook


def clone_workbook(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


def shrink_data_validations(ws):
    max_row = ws.max_row
    if max_row < 2:
        max_row = 2
    try:
        if hasattr(ws, 'data_validations') and hasattr(ws.data_validations, 'dataValidation'):
            for dv in ws.data_validations.dataValidation:
                if not dv.sqref:
                    continue
                ranges = str(dv.sqref).split()
                new_ranges = []
                for r in ranges:
                    if ':' in r:
                        start, end = r.split(':')
                        import re
                        match = re.match(r"([a-zA-Z]+)(\d+)", start)
                        if match:
                            col = match.group(1)
                            new_ranges.append(f"{start}:{col}{max_row}")
                        else:
                            new_ranges.append(r)
                    else:
                        new_ranges.append(r)
                dv.sqref = ' '.join(new_ranges)
    except Exception as e:
        print("Data validation shrink error:", e)

@app.route("/")
def index():
    from input_files import load_coupon_configs, load_net_discount_config
    return render_template(
        "index.html",
        report_columns=REPORT_COLUMNS,
        input_specs=INPUT_SPECS,
        uploaded_inputs=load_upload_status(UPLOAD_DIR, INPUT_MANIFEST),
        counter_configs=load_counter_configs(INPUT_MANIFEST),
        plus_extra_configs=load_plus_extra_configs(INPUT_MANIFEST),
        coupon_configs=load_coupon_configs(INPUT_MANIFEST),
        net_discount_config=load_net_discount_config(INPUT_MANIFEST),
        recommendation_rule=load_recommendation_rule(INPUT_MANIFEST),
    )


@app.errorhandler(413)
def upload_too_large(_error):
    return jsonify({"success": False, "message": "Yükleme toplam boyut sınırını aşıyor."}), 413


@app.route("/api/save-expiry", methods=["POST"])
def save_expiry():
    """Tarih bilgilerini hesaplama yapmadan anında kaydeder."""
    data = request.get_json(silent=True) or {}
    try:
        # Tek dosya expiry'leri
        single_expiries = data.get("single_expiries", {})
        if single_expiries and isinstance(single_expiries, dict):
            save_single_file_expiries(INPUT_MANIFEST, single_expiries)

        # Çoklu dosya (counter) expiry'leri
        counter_expiries = data.get("counter_expiries", {})
        if counter_expiries and isinstance(counter_expiries, dict):
            from input_files import load_counter_configs, save_counter_configs
            configs = load_counter_configs(INPUT_MANIFEST)
            for cfg in configs:
                cid = cfg.get("id", "")
                if cid in counter_expiries:
                    cfg["expiry_date"] = counter_expiries[cid]
            save_counter_configs(INPUT_MANIFEST, configs)

        # Çoklu dosya (plus_extra) expiry'leri
        plus_extra_expiries = data.get("plus_extra_expiries", {})
        if plus_extra_expiries and isinstance(plus_extra_expiries, dict):
            from input_files import load_plus_extra_configs, save_plus_extra_configs
            configs = load_plus_extra_configs(INPUT_MANIFEST)
            for cfg in configs:
                cid = cfg.get("id", "")
                if cid in plus_extra_expiries:
                    cfg["expiry_date"] = plus_extra_expiries[cid]
            save_plus_extra_configs(INPUT_MANIFEST, configs)

        # Çoklu dosya (coupon) expiry'leri
        coupon_expiries = data.get("coupon_expiries", {})
        if coupon_expiries and isinstance(coupon_expiries, dict):
            from input_files import load_coupon_configs, save_coupon_configs
            configs = load_coupon_configs(INPUT_MANIFEST)
            for cfg in configs:
                cid = cfg.get("id", "")
                if cid in coupon_expiries:
                    cfg["expiry_date"] = coupon_expiries[cid]
            save_coupon_configs(INPUT_MANIFEST, configs)

        # Net İndirim expiry
        net_discount_expiry = data.get("net_discount_expiry")
        if net_discount_expiry is not None:
            from input_files import load_net_discount_config, save_net_discount_config
            nd_cfg = load_net_discount_config(INPUT_MANIFEST)
            if nd_cfg:
                nd_cfg["expiry_date"] = net_discount_expiry
                save_net_discount_config(INPUT_MANIFEST, nd_cfg)

        return jsonify({"success": True})
    except Exception:
        app.logger.exception("Tarih kaydedilirken hata")
        return jsonify({"success": False, "message": "Tarih kaydedilemedi."}), 500


@app.route("/api/toggle-campaign-enabled", methods=["POST"])
def toggle_campaign_enabled():
    """Kampanya kartının Aktif/Pasif durumunu anında kaydedip saklar."""
    data = request.get_json(silent=True) or {}
    try:
        camp_type = data.get("type")  # 'counter', 'plus_extra', 'coupon', or 'net_discount'
        item_id = data.get("id")
        enabled = bool(data.get("enabled", True))

        if camp_type == "counter":
            from input_files import load_counter_configs, save_counter_configs
            configs = load_counter_configs(INPUT_MANIFEST)
            for item in configs:
                if item.get("id") == item_id:
                    item["enabled"] = enabled
            save_counter_configs(INPUT_MANIFEST, configs)
        elif camp_type == "plus_extra":
            from input_files import load_plus_extra_configs, save_plus_extra_configs
            configs = load_plus_extra_configs(INPUT_MANIFEST)
            for item in configs:
                if item.get("id") == item_id:
                    item["enabled"] = enabled
            save_plus_extra_configs(INPUT_MANIFEST, configs)
        elif camp_type == "coupon":
            from input_files import load_coupon_configs, save_coupon_configs
            configs = load_coupon_configs(INPUT_MANIFEST)
            for item in configs:
                if item.get("id") == item_id:
                    item["enabled"] = enabled
            save_coupon_configs(INPUT_MANIFEST, configs)
        elif camp_type == "net_discount":
            from input_files import load_net_discount_config, save_net_discount_config
            nd_cfg = load_net_discount_config(INPUT_MANIFEST)
            if nd_cfg:
                nd_cfg["enabled"] = enabled
                save_net_discount_config(INPUT_MANIFEST, nd_cfg)
        elif camp_type == "single_file" or item_id in INPUT_SPECS:
            from input_files import save_single_file_enabled
            save_single_file_enabled(INPUT_MANIFEST, item_id, enabled)

        return jsonify({"success": True})
    except Exception as e:
        app.logger.exception("Kampanya durumu kaydedilirken hata")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove-net-discount-file", methods=["POST"])
def remove_net_discount_file():
    try:
        from input_files import load_net_discount_config, save_net_discount_config
        nd_cfg = load_net_discount_config(INPUT_MANIFEST)
        if nd_cfg:
            p = nd_cfg.get("path") or nd_cfg.get("stored_path")
            if p and os.path.exists(p):
                try: os.remove(p)
                except Exception: pass
            save_net_discount_config(INPUT_MANIFEST, {})
        return jsonify({"success": True, "net_discount_config": {}})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove-counter-file", methods=["POST"])
def remove_counter_file():
    try:
        data = request.get_json(silent=True) or {}
        item_id = data.get("id")
        item_path = data.get("path")
        
        counter_configs = load_counter_configs(INPUT_MANIFEST)
        new_configs = []
        for item in counter_configs:
            p = item.get("path") or item.get("stored_path")
            if (item_id and item.get("id") == item_id) or (item_path and p == item_path):
                if p and os.path.exists(p):
                    try: os.remove(p)
                    except Exception: pass
            else:
                new_configs.append(item)
        save_counter_configs(INPUT_MANIFEST, new_configs)
        return jsonify({"success": True, "counter_configs": new_configs})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove-plus-extra-file", methods=["POST"])
def remove_plus_extra_file():
    try:
        data = request.get_json(silent=True) or {}
        item_id = data.get("id")
        item_path = data.get("path")
        
        plus_extra_configs = load_plus_extra_configs(INPUT_MANIFEST)
        new_configs = []
        for item in plus_extra_configs:
            p = item.get("path") or item.get("stored_path")
            if (item_id and item.get("id") == item_id) or (item_path and p == item_path):
                if p and os.path.exists(p):
                    try: os.remove(p)
                    except Exception: pass
            else:
                new_configs.append(item)
        save_plus_extra_configs(INPUT_MANIFEST, new_configs)
        return jsonify({"success": True, "plus_extra_configs": new_configs})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/remove-coupon-file", methods=["POST"])
def remove_coupon_file():
    try:
        data = request.get_json(silent=True) or {}
        item_id = data.get("id")
        item_path = data.get("path")
        
        from input_files import load_coupon_configs, save_coupon_configs
        coupon_configs = load_coupon_configs(INPUT_MANIFEST)
        new_configs = []
        for item in coupon_configs:
            p = item.get("path") or item.get("stored_path")
            if (item_id and item.get("id") == item_id) or (item_path and p == item_path):
                if p and os.path.exists(p):
                    try: os.remove(p)
                    except Exception: pass
            else:
                new_configs.append(item)
        save_coupon_configs(INPUT_MANIFEST, new_configs)
        return jsonify({"success": True, "coupon_configs": new_configs})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/data")
def get_data():
    if not calculation_result_is_current():
        return jsonify({"needs_calculation": True, "message": "Lütfen önce 'Verileri Güncelle' butonuna basarak hesaplamaları başlatın."}), 200
        
    df = restore_persisted_collections(pd.read_excel(F_HESAP))
    records = df.to_dict(orient="records")
    user_selections = load_user_selections(INPUT_MANIFEST)
    if user_selections and isinstance(user_selections, dict):
        for rec in records:
            b = str(rec.get("Barkod", "")).strip()
            if b in user_selections:
                val = user_selections[b]
                if isinstance(val, dict):
                    rec["İlk Kampanya Seçimi"] = val.get("main", "Hiçbiri")
                    rec["İlk Ekstra Kampanya Seçimi"] = val.get("extra", "Hiçbiri")
                else:
                    rec["İlk Kampanya Seçimi"] = str(val)
    cleaned_records = clean_nans(records)
    return jsonify(cleaned_records), 200

@app.route("/api/save-selections", methods=["POST"])
def save_selections_endpoint():
    try:
        data = request.get_json(silent=True)
        selections = data.get("selections") if isinstance(data, dict) else None
        if not selection_payload_is_valid(selections):
            return jsonify({"success": False, "message": "Geçersiz veri biçimi."}), 400
        normalized_selections = {}
        for barcode, selection in selections.items():
            main, extra = normalize_selection(selection)
            normalized_selections[barcode] = {"main": main, "extra": extra}
        save_user_selections(INPUT_MANIFEST, normalized_selections)
        return jsonify({"success": True, "saved_count": len(normalized_selections)})
    except Exception:
        app.logger.exception("Seçimler kaydedilirken hata")
        return jsonify({"success": False, "message": "Seçimler kaydedilemedi."}), 500


@app.route("/api/recommendation-rule", methods=["POST"])
def save_recommendation_rule_endpoint():
    try:
        payload = request.get_json(silent=True)
        if payload is None:
            raise InputValidationError("Öneri kuralı boş olamaz.")
        if isinstance(payload, dict) and "recommendation_rule" in payload:
            if set(payload) != {"recommendation_rule"}:
                raise InputValidationError("Öneri kuralı isteği geçersiz.")
            payload = payload["recommendation_rule"]
        rule = save_recommendation_rule(INPUT_MANIFEST, payload)
        return jsonify({"success": True, "recommendation_rule": rule})
    except InputValidationError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception:
        app.logger.exception("Öneri kuralı kaydedilirken hata")
        return jsonify({"success": False, "message": "Öneri kuralı kaydedilemedi."}), 500

@app.route("/api/calculate", methods=["POST"])
def calculate():
    try:
        from komisyon_hesaplayici import calculate_all
        from input_files import save_counter_configs, INPUT_SPECS

        counter_configs = parse_campaign_configs_json(
            request.form.get("counter_configs_json"), "counter"
        )
        plus_extra_configs = parse_campaign_configs_json(
            request.form.get("plus_extra_configs_json"), "plus_extra"
        )
        coupon_configs = parse_campaign_configs_json(
            request.form.get("coupon_configs_json"), "coupon"
        )
        recommendation_rule_raw = request.form.get("recommendation_rule_json")
        recommendation_rule = (
            load_recommendation_rule(INPUT_MANIFEST)
            if recommendation_rule_raw is None
            else parse_recommendation_rule_json(recommendation_rule_raw)
        )

        standard_files = {
            k: v for k, v in request.files.items() 
            if k in INPUT_SPECS and v and v.filename
        }

        input_files = save_upload_set(standard_files, UPLOAD_DIR, INPUT_MANIFEST)
        missing_base_inputs = [
            INPUT_SPECS[key]["label"]
            for key in ("discount", "commission", "current")
            if not input_files.get(key)
        ]
        if missing_base_inputs:
            raise InputValidationError(
                f"Zorunlu girdiler eksik: {', '.join(missing_base_inputs)}"
            )
        if recommendation_rule_raw is not None:
            save_recommendation_rule(INPUT_MANIFEST, recommendation_rule)

        single_expiries_raw = request.form.get("single_expiries_json")
        if single_expiries_raw:
            try:
                single_expiries = json.loads(single_expiries_raw)
                save_single_file_expiries(INPUT_MANIFEST, single_expiries)
            except Exception as exp_err:
                print("Single expiries parse error:", exp_err)

        single_enabled_raw = request.form.get("single_enabled_json")
        if single_enabled_raw:
            try:
                single_enabled = json.loads(single_enabled_raw)
                if isinstance(single_enabled, dict):
                    from input_files import save_single_file_enabled
                    for k, v in single_enabled.items():
                        if k in INPUT_SPECS:
                            save_single_file_enabled(INPUT_MANIFEST, k, bool(v))
            except Exception as en_err:
                print("Single enabled parse error:", en_err)

        counter_files = []
        counter_dir = os.path.join(UPLOAD_DIR, "counter_files")
        os.makedirs(counter_dir, exist_ok=True)

        for idx, item in enumerate(counter_configs):
            file_key = f"counter_file_{idx}"
            file_obj = request.files.get(file_key)
            stored_name = f"counter_{idx+1}.xlsx"
            target_path = os.path.join(counter_dir, stored_name)
            stored_path = item.get("stored_path") or item.get("path")
            original_name = item.get("original_name")

            if file_obj and file_obj.filename:
                file_obj.save(target_path)
                stored_path = target_path
                original_name = file_obj.filename
            elif os.path.exists(target_path):
                stored_path = target_path

            counter_files.append({
                **item,
                "id": item.get("id", f"counter_{idx+1}"),
                "label": build_campaign_label(item, "counter", idx),
                "filename": item.get("filename") or original_name or stored_name,
                "original_name": original_name,
                "path": stored_path,
                "stored_path": stored_path,
                "expiry_date": item.get("expiry_date", ""),
                "enabled": item.get("enabled", True) is not False,
            })

        save_counter_configs(INPUT_MANIFEST, counter_files)

        from input_files import save_plus_extra_configs, load_plus_extra_configs

        plus_extra_files = []
        plus_extra_dir = os.path.join(UPLOAD_DIR, "plus_extra_files")
        os.makedirs(plus_extra_dir, exist_ok=True)

        for idx, item in enumerate(plus_extra_configs):
            file_key = f"plus_extra_file_{idx}"
            file_obj = request.files.get(file_key)
            stored_name = f"plus_extra_{idx+1}.xlsx"
            target_path = os.path.join(plus_extra_dir, stored_name)
            stored_path = item.get("stored_path") or item.get("path")
            original_name = item.get("original_name")

            if file_obj and file_obj.filename:
                file_obj.save(target_path)
                stored_path = target_path
                original_name = file_obj.filename
            elif os.path.exists(target_path):
                stored_path = target_path

            plus_extra_files.append({
                **item,
                "id": item.get("id", f"plus_extra_{idx+1}"),
                "label": build_campaign_label(item, "plus_extra", idx),
                "filename": item.get("filename") or original_name or stored_name,
                "original_name": original_name,
                "path": stored_path,
                "stored_path": stored_path,
                "expiry_date": item.get("expiry_date", ""),
                "enabled": item.get("enabled", True) is not False,
            })

        save_plus_extra_configs(INPUT_MANIFEST, plus_extra_files)

        from input_files import save_coupon_configs, load_coupon_configs

        coupon_files = []
        coupon_dir = os.path.join(UPLOAD_DIR, "coupon_files")
        os.makedirs(coupon_dir, exist_ok=True)

        for idx, item in enumerate(coupon_configs):
            file_key = f"coupon_file_{idx}"
            file_obj = request.files.get(file_key)
            stored_name = f"coupon_{idx+1}.xlsx"
            target_path = os.path.join(coupon_dir, stored_name)
            stored_path = item.get("stored_path") or item.get("path")
            original_name = item.get("original_name")

            if file_obj and file_obj.filename:
                file_obj.save(target_path)
                stored_path = target_path
                original_name = file_obj.filename
            elif os.path.exists(target_path):
                stored_path = target_path

            coupon_files.append({
                **item,
                "id": item.get("id", f"coupon_{idx+1}"),
                "label": build_campaign_label(item, "coupon", idx),
                "filename": item.get("filename") or original_name or stored_name,
                "original_name": original_name,
                "path": stored_path,
                "stored_path": stored_path,
                "expiry_date": item.get("expiry_date", ""),
                "enabled": item.get("enabled", True) is not False,
            })

        save_coupon_configs(INPUT_MANIFEST, coupon_files)

        from input_files import save_net_discount_config, load_net_discount_config
        nd_config_raw = request.form.get("net_discount_config_json")
        net_discount_config = None
        if nd_config_raw:
            try:
                parsed_nd = json.loads(nd_config_raw)
                if isinstance(parsed_nd, dict) and parsed_nd:
                    net_discount_config = normalize_campaign_config(parsed_nd, "net_discount")
            except Exception:
                pass
        if net_discount_config is None:
            net_discount_config = load_net_discount_config(INPUT_MANIFEST) or {}

        nd_file = request.files.get("net_discount_file") or request.files.get("net_discount")
        nd_path = os.path.join(UPLOAD_DIR, "net_discount.xlsx")
        if nd_file and nd_file.filename:
            nd_file.save(nd_path)
            if not net_discount_config:
                net_discount_config = {"discount_type": "%", "discount_amount": 0.0}
            net_discount_config["path"] = nd_path
            net_discount_config["stored_path"] = nd_path
            net_discount_config["original_name"] = nd_file.filename
            net_discount_config["filename"] = nd_file.filename
        elif os.path.exists(nd_path) and net_discount_config:
            net_discount_config["path"] = nd_path
            net_discount_config["stored_path"] = nd_path

        if net_discount_config:
            net_discount_config["label"] = build_campaign_label(net_discount_config, "net_discount")
            save_net_discount_config(INPUT_MANIFEST, net_discount_config)

        toplam_indirim = float(request.form.get("toplam_indirim", 0) or 0)
        trendyol_oran = float(request.form.get("trendyol_oran", 0) or 0)
        min_sepet = float(request.form.get("min_sepet", 0) or 0)
        karsilamali_config = {
            "min_sepet": min_sepet,
            "toplam_indirim": toplam_indirim,
            "trendyol_oran": trendyol_oran,
        }

        # Reset saved user selections on new calculation so all products start fresh as 'Hiçbiri'
        save_user_selections(INPUT_MANIFEST, {})
        user_selections = {}

        # Sadece Aktif olan tekil dosyaları hesaplamaya dahil et (Pasif olanları muaf tut)
        upload_status = load_upload_status(UPLOAD_DIR, INPUT_MANIFEST)
        active_input_files = {}
        for k, p in input_files.items():
            st = upload_status.get(k, {})
            if st.get("enabled", True) is not False:
                active_input_files[k] = p
            else:
                active_input_files[k] = None

        result = calculate_all(active_input_files, counter_files=counter_files, plus_extra_files=plus_extra_files, coupon_files=coupon_files, net_discount_config=net_discount_config, karsilamali_config=karsilamali_config, output_dir=OUTPUT_DIR, user_selections=user_selections, recommendation_rule=recommendation_rule)
        if result.get("success"):
            result["uploads"] = load_upload_status(UPLOAD_DIR, INPUT_MANIFEST)
            result["counter_configs"] = load_counter_configs(INPUT_MANIFEST)
            result["plus_extra_configs"] = load_plus_extra_configs(INPUT_MANIFEST)
            result["coupon_configs"] = load_coupon_configs(INPUT_MANIFEST)
            result["net_discount_config"] = load_net_discount_config(INPUT_MANIFEST)
            result["recommendation_rule"] = recommendation_rule
            pd.DataFrame(result["results"]).to_excel(F_HESAP, index=False)
            result = clean_nans(result)
        return jsonify(result), (200 if result.get("success") else 500)
    except InputValidationError as e:
        return jsonify({"success": False, "message": str(e)}), 400
    except ValueError:
        return jsonify({"success": False, "message": "Sayısal kampanya değerlerini kontrol edin."}), 400
    except Exception:
        app.logger.exception("Hesaplama sırasında hata")
        return jsonify({"success": False, "message": "Hesaplama sırasında beklenmeyen bir hata oluştu."}), 500


@app.route("/api/apply", methods=["POST"])
def apply_campaign():
    data = request.get_json(silent=True) or {}
    selections = data.get("selections", {})
    if not selection_payload_is_valid(selections):
        return jsonify({"success": False, "message": "Kampanya seçimleri geçersiz."}), 400
    visible_columns = data.get("visibleColumns")
    target_type = data.get("target_type", "Hepsi")
    if target_type not in VALID_TARGET_TYPES:
        return jsonify({"success": False, "message": "Çıktı türü geçersiz."}), 400
    try:
        input_files = load_upload_set(UPLOAD_DIR, INPUT_MANIFEST)
    except InputValidationError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    if not calculation_result_is_current():
        return jsonify({
            "success": False,
            "message": "Önce yüklenen girdilerle hesaplama yapın.",
        }), 400
    try:
        df_hesap = pd.read_excel(F_HESAP)
        if not REQUIRED_RESULT_COLUMNS.issubset(df_hesap.columns):
            return jsonify({
                "success": False,
                "message": "Önce yüklenen girdilerle hesaplama yapın.",
            }), 400
        raw_rows = restore_persisted_collections(
            df_hesap
        ).to_dict(orient="records")
    except Exception:
        app.logger.exception("Hesap sonucu okunamadı")
        return jsonify({
            "success": False,
            "message": "Hesap sonucu okunamadı; girdileri yeniden hesaplayın.",
        }), 500

    ignore_zero_stock = bool(data.get("ignore_zero_stock", True))

    table_data = []
    for row in raw_rows:
        barcode = str(row.get("Barkod", "")).strip()
        stok = as_number(row.get("Stok Adedi"))
        if ignore_zero_stock and stok is not None and stok == 0:
            continue

        selection = selections.get(barcode, {
            "main": row.get("İlk Kampanya Seçimi", "Hiçbiri") or "Hiçbiri",
            "extra": row.get("İlk Ekstra Kampanya Seçimi", "Hiçbiri") or "Hiçbiri",
        })
        normalized = normalize_selection(selection)
        if not normalized:
            return jsonify({
                "success": False,
                "message": "Kampanya seçimleri geçersiz.",
            }), 400
        main_selection, extra_selection = normalized
        if not row_selection_is_applicable(row, main_selection, extra_selection) or not selection_pair_is_safe(row, main_selection, extra_selection):
            return jsonify({
                "success": False,
                "message": "Seçilen kampanya ürün için uygulanabilir değil.",
            }), 400
        row["userSelection"] = main_selection
        row["userExtraSelection"] = extra_selection
        table_data.append(row)

    F_AVAN = input_files.get("advantage")
    F_FLAS = input_files.get("flash")
    F_PLUS = input_files.get("plus")
    F_PLUS_EK = input_files.get("plus_extra")
    F_KARS = input_files.get("counter")
    F_MUH_AVAN = input_files.get("muhasebe_avantaj")
    F_MUH_FLAS = input_files.get("muhasebe_flas")
    F_MUH_PLUS = input_files.get("muhasebe_plus")
    try:
        plus_extra_configs = load_plus_extra_configs(INPUT_MANIFEST)
        from input_files import load_net_discount_config
        nd_config = load_net_discount_config(INPUT_MANIFEST)
    except InputValidationError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    target_inputs = {
        "Avantajlı": F_AVAN or F_MUH_AVAN,
        "Flaş": F_FLAS or F_MUH_FLAS,
        "Plus": F_PLUS or F_MUH_PLUS,
        "Plus Ek İndirim": F_PLUS_EK or any(
            config.get("path") or config.get("stored_path")
            for config in plus_extra_configs
        ),
        "Karşılamalı Kampanya": F_KARS,
    }
    if target_type != "Hepsi" and target_type in target_inputs and not target_inputs[target_type]:
        return jsonify({
            "success": False,
            "message": f"{target_type} girdisi bu hesaplamada yüklenmedi.",
        }), 400
        
    # Her işlem için Tarih_Saat adında alt klasör oluştur
    timestamp_folder = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_output_dir = os.path.join(OUTPUT_DIR, timestamp_folder)
    os.makedirs(run_output_dir, exist_ok=True)
    generated_files = []

    # Map barcodes to their table_data row dict for easy lookup of match statuses & recommendations
    row_by_barcode = {}
    for row in table_data:
        b_key = str(row.get("Barkod", "")).strip()
        if b_key:
            row_by_barcode[b_key] = row

    def get_selection(b_key):
        row = row_by_barcode.get(b_key, {})
        return (
            row.get("userSelection", "Hiçbiri") or "Hiçbiri",
            row.get("userExtraSelection", "Hiçbiri") or "Hiçbiri",
        )

    # 1. Process Avantajlı
    if target_type in ['Hepsi', 'Avantajlı'] and (F_AVAN or F_MUH_AVAN):
        try:
            wb_av = load_merged_campaign_workbook(F_AVAN, F_MUH_AVAN, "BARKOD")
            ws_av = wb_av.active
            b_idx_av = header_index(ws_av, 'BARKOD')
            tsf_idx_av = ensure_header(ws_av, 'YENİ TSF (FİYAT GÜNCELLE)')
            tarife_idx_av = ensure_header(ws_av, 'Tarife Sonuna Kadar Uygula')
        
            if b_idx_av and tsf_idx_av:
                keep_rows = []
                for r in range(2, ws_av.max_row + 1):
                    b_val = ws_av.cell(r, b_idx_av).value
                    if not b_val: continue
                    b_val_str = str(b_val).strip()
                    main_sel, extra_sel = get_selection(b_val_str)
                    row_info = row_by_barcode.get(b_val_str, {})
                    should_keep = (main_sel == "Avantajlı")

                    if should_keep:
                        selected_price = row_info.get('Avantajlı Ürün Fiyatı (YENİ TSF) (TL)')
                        if selected_price is not None and not pd.isna(selected_price):
                            ws_av.cell(r, tsf_idx_av).value = float(selected_price)
                        if tarife_idx_av:
                            ws_av.cell(r, tarife_idx_av).value = "Evet"
                        keep_rows.append(r)
                
                if keep_rows:
                    safe_keep_rows(ws_av, keep_rows)
                    out_name = os.path.join(run_output_dir, "Avantajli_Urun_Etiketleri.xlsx")
                    shrink_data_validations(ws_av)
                    wb_av.save(out_name)
                    fix_xlsx_for_trendyol(out_name)
                    generated_files.append(os.path.join(timestamp_folder, "Avantajli_Urun_Etiketleri.xlsx"))
        except Exception:
            return processing_error("Avantajlı dosya")

    # 2. Process Flaş (Grouped by Date)
    if target_type in ['Hepsi', 'Flaş'] and (F_FLAS or F_MUH_FLAS):
        try:
            wb_fl = load_merged_campaign_workbook(F_FLAS, F_MUH_FLAS, "Barkod")
            ws_fl = wb_fl.active
            b_idx_fl = header_index(ws_fl, 'Barkod')
            fiyat_24_idx = ensure_header(ws_fl, '24 Saat Fiyat')
            fiyat_3_idx = header_index(ws_fl, '3 Saat Fiyat')
            fixed_price_idx = ensure_header(ws_fl, 'Senin Belirlediğin Flaş Fiyatı')
            guncel_fiyat_idx = ensure_header(ws_fl, 'Güncellenecek Fiyat')
            baslangic_idx = ensure_header(ws_fl, '24 Saat Flaş Başlangıç Tarihi')
            bitis_idx = header_index(ws_fl, '24 Saat Flaş Bitiş Tarihi')
            baslangic_3_idx = header_index(ws_fl, '3 Saat Flaş Başlangıç Tarihi')
            bitis_3_idx = header_index(ws_fl, '3 Saat Flaş Bitiş Tarihi')

            if b_idx_fl and fiyat_24_idx and guncel_fiyat_idx and baslangic_idx:
                period_columns = {
                    "24 Saat": {
                        "price": fiyat_24_idx,
                        "start": baslangic_idx,
                        "end": bitis_idx,
                    },
                    "3 Saat": {
                        "price": fiyat_3_idx,
                        "start": baslangic_3_idx,
                        "end": bitis_3_idx,
                    },
                }

                def source_intervals(row_number):
                    intervals = []
                    for period, columns in period_columns.items():
                        if not columns["price"]:
                            continue
                        start = (
                            ws_fl.cell(row_number, columns["start"]).value
                            if columns["start"]
                            else None
                        )
                        end = (
                            ws_fl.cell(row_number, columns["end"]).value
                            if columns["end"]
                            else None
                        )
                        period_price = ws_fl.cell(row_number, columns["price"]).value
                        if (
                            period_price not in (None, "")
                            or start not in (None, "")
                            or end not in (None, "")
                        ):
                            intervals.append((period, start, end, columns["price"]))
                    if not intervals:
                        intervals.append(("24 Saat", None, None, fiyat_24_idx))
                    return intervals

                source_interval_counts = {}
                source_interval_keys = {}
                for r in range(2, ws_fl.max_row + 1):
                    barcode = str(ws_fl.cell(r, b_idx_fl).value or "").strip()
                    if barcode:
                        intervals = source_intervals(r)
                        source_interval_counts[barcode] = (
                            source_interval_counts.get(barcode, 0)
                            + len(intervals)
                        )
                        source_interval_keys.setdefault(barcode, []).extend(
                            (
                                period,
                                normalize_flash_interval_value(start),
                                normalize_flash_interval_value(end),
                            )
                            for period, start, end, _price_column in intervals
                        )

                date_groups = {}  # {(date_str, period): [(row, price, column, label)]}
                ambiguous_legacy = set()

                for r in range(2, ws_fl.max_row + 1):
                    b_val = ws_fl.cell(r, b_idx_fl).value
                    if not b_val: continue

                    b_val_str = str(b_val).strip()
                    main_sel, extra_sel = get_selection(b_val_str)
                    should_keep = (main_sel == "Flaş")

                    if should_keep:
                        row_info = row_by_barcode.get(b_val_str, {})
                        has_interval_contract = "flash_evaluations" in row_info
                        evaluations = row_info.get("flash_evaluations", [])
                        if not isinstance(evaluations, list):
                            evaluations = []

                        ambiguous_fixed = any(
                            isinstance(item, dict)
                            and item.get("eligible") is True
                            and str(item.get("source") or "").strip().casefold()
                            == "senin belirlediğin flaş fiyatı".casefold()
                            and not normalize_flash_interval_value(item.get("start"))
                            and not normalize_flash_interval_value(item.get("end"))
                            and (
                                normalize_flash_period(item.get("period")),
                                "",
                                "",
                            ) not in source_interval_keys.get(b_val_str, [])
                            and not (
                                source_interval_counts.get(b_val_str) == 1
                                and len(evaluations) == 1
                            )
                            for item in evaluations
                        )
                        if has_interval_contract and ambiguous_fixed:
                            ambiguous_legacy.add(b_val_str)
                            continue

                        if not has_interval_contract and source_interval_counts.get(b_val_str) != 1:
                            ambiguous_legacy.add(b_val_str)
                            continue

                        for period, start, end, period_price_idx in source_intervals(r):
                            evaluation = None
                            if has_interval_contract:
                                interval_key = (
                                    period,
                                    normalize_flash_interval_value(start),
                                    normalize_flash_interval_value(end),
                                )
                                evaluation = next((
                                    item
                                    for item in evaluations
                                    if isinstance(item, dict)
                                    and (
                                        normalize_flash_period(item.get("period")),
                                        normalize_flash_interval_value(item.get("start")),
                                        normalize_flash_interval_value(item.get("end")),
                                    ) == interval_key
                                ), None)
                                if (
                                    evaluation is None
                                    and source_interval_counts.get(b_val_str) == 1
                                    and len(evaluations) == 1
                                ):
                                    fixed_evaluation = evaluations[0]
                                    if (
                                        isinstance(fixed_evaluation, dict)
                                        and normalize_flash_period(
                                            fixed_evaluation.get("period")
                                        ) == period
                                        and not normalize_flash_interval_value(
                                            fixed_evaluation.get("start")
                                        )
                                        and not normalize_flash_interval_value(
                                            fixed_evaluation.get("end")
                                        )
                                        and str(
                                            fixed_evaluation.get("source") or ""
                                        ).strip().casefold()
                                        == "senin belirlediğin flaş fiyatı".casefold()
                                    ):
                                        evaluation = fixed_evaluation
                                if not evaluation or evaluation.get("eligible") is not True:
                                    continue
                                selected_price = as_number(evaluation.get("price"))
                                source_name = str(evaluation.get("source") or "").strip()
                                source_period = normalize_flash_period(source_name)
                                if source_name.casefold() == "senin belirlediğin flaş fiyatı".casefold():
                                    if not fixed_price_idx:
                                        continue
                                    target_price_idx = fixed_price_idx
                                    selection_label = "Senin Belirlediğin Flaş Fiyatı"
                                elif source_name.casefold() == "mevcut fiyat":
                                    target_price_idx = period_price_idx
                                    selection_label = period
                                elif source_period == period:
                                    target_price_idx = period_price_idx
                                    selection_label = period
                                else:
                                    continue
                            else:
                                selected_price = as_number(
                                    row_info.get('Flaş Ürün 24 Saat Fiyatı (TL)')
                                )
                                target_price_idx = period_price_idx
                                selection_label = period

                            if selected_price is None:
                                continue
                            date_val = str(start or "").strip()
                            date_key = (
                                date_val.split()[0].replace('/', '_').replace('-', '_')
                                if date_val
                                else "Genel"
                            )
                            date_groups.setdefault((date_key, period), []).append((
                                r,
                                selected_price,
                                target_price_idx,
                                selection_label,
                            ))

                if ambiguous_legacy:
                    return jsonify({
                        "success": False,
                        "message": (
                            "Flaş sonuçları tarih aralıklarını ayırt etmiyor; "
                            "yüklenen girdilerle yeniden hesaplayın; yeni hesaplama yapın."
                        ),
                    }), 400

                for (date_key, period), assignments in date_groups.items():
                    if assignments:
                        wb_copy = clone_workbook(wb_fl)
                        ws_copy = wb_copy.active
                        for r, selected_price, target_price_idx, selection_label in assignments:
                            ws_copy.cell(r, target_price_idx).value = float(selected_price)
                            ws_copy.cell(r, guncel_fiyat_idx).value = selection_label
                        keep_rows = [assignment[0] for assignment in assignments]
                        safe_keep_rows(ws_copy, keep_rows)
                        period_suffix = "_3_Saat" if period == "3 Saat" else ""
                        file_name = f"Flas_Urunler_{date_key}{period_suffix}.xlsx"
                        out_name = os.path.join(run_output_dir, file_name)
                        shrink_data_validations(ws_copy)
                        wb_copy.save(out_name)
                        fix_xlsx_for_trendyol(out_name)
                        generated_files.append(os.path.join(timestamp_folder, file_name))
                
        except Exception:
            return processing_error("Flaş dosya")

    # 3. Process Plus (Grouped by Date Interval)
    if target_type in ['Hepsi', 'Plus']:
        if F_PLUS or F_MUH_PLUS:
            try:
                wb_plus = load_merged_campaign_workbook(F_PLUS, F_MUH_PLUS, "Barkod")
                ws_plus = wb_plus.active
                b_idx_plus = header_index(ws_plus, 'Barkod')
                fiyat_secim_idx = ensure_header(ws_plus, 'Plus Fiyat Seçimi')
                tarife_secim_idx = ensure_header(ws_plus, 'Tarife Seçimi')
                ust_limit_idx = ensure_header(ws_plus, 'Plus Fiyat Üst Limiti')
                header_plus = [ws_plus.cell(1, c).value for c in range(1, ws_plus.max_column + 1)]
                plus_periods = find_plus_period_columns(header_plus)
                legacy_tariff = f"{sum(period['days'] for period in plus_periods) or 7} Günlük Fiyat"

                if b_idx_plus and fiyat_secim_idx and tarife_secim_idx and ust_limit_idx:
                    date_groups = {}  # {all_period_values: [row_indices]}

                    for r in range(2, ws_plus.max_row + 1):
                        b_val = ws_plus.cell(r, b_idx_plus).value
                        if not b_val: continue
                        b_val_str = str(b_val).strip()
                        main_sel, extra_sel = get_selection(b_val_str)
                        should_keep = (main_sel == "Plus")

                        if should_keep:
                            period_values = tuple(
                                ws_plus.cell(r, period["date_position"] + 1).value
                                for period in plus_periods
                            )
                            date_groups.setdefault(period_values, []).append(r)

                    used_file_names = set()
                    for period_values, keep_rows in date_groups.items():
                        if keep_rows:
                            wb_copy = clone_workbook(wb_plus)
                            ws_copy = wb_copy.active
                            for r in keep_rows:
                                ust_lim = ws_copy.cell(r, ust_limit_idx).value
                                barcode = str(ws_copy.cell(r, b_idx_plus).value or '').strip()
                                row_info = row_by_barcode.get(barcode, {})
                                selected_price = row_info.get('Plus Fiyatı (TL)')
                                ws_copy.cell(r, fiyat_secim_idx).value = selected_price if selected_price is not None and not pd.isna(selected_price) else ust_lim
                                selected_tariff = row_info.get('Plus Tarife Seçimi')
                                ws_copy.cell(r, tarife_secim_idx).value = (
                                    selected_tariff
                                    if isinstance(selected_tariff, str) and selected_tariff.strip()
                                    else legacy_tariff
                                )

                            safe_keep_rows(ws_copy, keep_rows)
                            for table in ws_copy.tables.values():
                                min_col, min_row, max_col, max_row = (
                                    openpyxl.utils.range_boundaries(table.ref)
                                )
                                if max_row > ws_copy.max_row:
                                    table.ref = (
                                        f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
                                        f"{openpyxl.utils.get_column_letter(max_col)}{ws_copy.max_row}"
                                    )
                            date_parts = []
                            for position, value in enumerate(period_values, 1):
                                clean_value = str(value or "").strip()
                                if clean_value:
                                    clean_value = clean_value.translate(
                                        str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
                                    )
                                    clean_value = re.sub(r'[^\w\.\-]', '_', clean_value)
                                    clean_value = re.sub(r'_+', '_', clean_value).strip('_')
                                    if clean_value:
                                        date_parts.append(
                                            f"{position}_{clean_value}"
                                            if len(plus_periods) > 1
                                            else clean_value
                                        )
                            date_key = "__".join(date_parts) or "Genel"
                            if date_key == "Genel":
                                file_name = "Plus_Komisyon_Tarifeleri.xlsx"
                            else:
                                file_name = f"Plus_Komisyon_Tarifeleri_{date_key}.xlsx"
                            if file_name in used_file_names:
                                base_name, extension = os.path.splitext(file_name)
                                suffix = 2
                                while f"{base_name}_{suffix}{extension}" in used_file_names:
                                    suffix += 1
                                file_name = f"{base_name}_{suffix}{extension}"
                            used_file_names.add(file_name)

                            out_name = os.path.join(run_output_dir, file_name)
                            shrink_data_validations(ws_copy)
                            wb_copy.save(out_name)
                            fix_xlsx_for_trendyol(out_name)
                            generated_files.append(os.path.join(timestamp_folder, file_name))
            except Exception:
                return processing_error("Plus dosya")

    # 3.5 Process Ürün Komisyon Tarifeleri
    if target_type in ['Hepsi', 'Komisyon Tarifesi']:
        kom_path = input_files.get('commission')
        if kom_path and os.path.exists(kom_path):
            try:
                wb_kom = openpyxl.load_workbook(kom_path)
                ws_kom = wb_kom.active
                b_idx_kom = header_index(ws_kom, 'Barkod') or header_index(ws_kom, 'BARKOD')
                fiyat_idx_kom = ensure_header(ws_kom, 'YENİ TSF (FİYAT GÜNCELLE)')
                secim_idx_kom = ensure_header(ws_kom, 'Tarife Seçimi')

                if b_idx_kom:
                    keep_rows = []
                    for r in range(2, ws_kom.max_row + 1):
                        b_val = ws_kom.cell(r, b_idx_kom).value
                        if not b_val:
                            continue
                        b_val_str = str(b_val).strip()
                        main_sel, extra_sel = get_selection(b_val_str)
                        if main_sel == "Komisyon Tarifesi":
                            row_info = row_by_barcode.get(b_val_str, {})
                            if fiyat_idx_kom:
                                sel_price = row_info.get('Komisyon Tarifesi Fiyatı (TL)')
                                if sel_price is not None and not pd.isna(sel_price):
                                    ws_kom.cell(r, fiyat_idx_kom).value = float(sel_price)
                            if secim_idx_kom:
                                sel_tarife = row_info.get('Komisyon Tarifesi Seçimi') or '7 Günlük Fiyat'
                                ws_kom.cell(r, secim_idx_kom).value = str(sel_tarife)
                            keep_rows.append(r)

                    if keep_rows:
                        safe_keep_rows(ws_kom, keep_rows)
                        file_name = "Urun_Komisyon_Tarifeleri_Urunler.xlsx"
                        out_name = os.path.join(run_output_dir, file_name)
                        shrink_data_validations(ws_kom)
                        wb_kom.save(out_name)
                        fix_xlsx_for_trendyol(out_name)
                        generated_files.append(os.path.join(timestamp_folder, file_name))
            except Exception as e:
                print("Komisyon Tarifesi export error:", e)
                return processing_error("Ürün Komisyon Tarifeleri dosya")

    # 4. Process Plus Ek İndirim (Çoklu Dosya Desteği)
    if target_type in ['Hepsi', 'Plus Ek İndirim']:
        try:
            plus_extra_configs = load_plus_extra_configs(INPUT_MANIFEST)

            for idx, pe_item in enumerate(plus_extra_configs):
                pe_item = normalize_campaign_config(pe_item, "plus_extra")
                pe_path = pe_item.get('stored_path') or pe_item.get('path')
                if pe_path and os.path.exists(pe_path):
                    wb_pe = openpyxl.load_workbook(pe_path)
                    ws_pe = wb_pe.active
                    header_pe = [ws_pe.cell(1, c).value for c in range(1, ws_pe.max_column + 1)]
                    b_idx_pe = header_pe.index('Barkod') + 1 if 'Barkod' in header_pe else None
                    fiyat_idx_pe = header_pe.index('Kampanyalı Satış Fiyatı') + 1 if 'Kampanyalı Satış Fiyatı' in header_pe else None
                    max_fiyat_idx_pe = header_pe.index('Maksimum Girebileceğin Fiyat') + 1 if 'Maksimum Girebileceğin Fiyat' in header_pe else None
                    
                    if b_idx_pe and fiyat_idx_pe and max_fiyat_idx_pe:
                        c_label = build_campaign_label(pe_item, "plus_extra", idx)
                        keep_rows = []
                        for r in range(2, ws_pe.max_row + 1):
                            b_val = ws_pe.cell(r, b_idx_pe).value
                            if not b_val: continue

                            b_val_str = str(b_val).strip()
                            main_sel, extra_sel = get_selection(b_val_str)
                            should_keep = extra_sel == c_label
                            if should_keep:
                                max_f = ws_pe.cell(r, max_fiyat_idx_pe).value
                                row_info = row_by_barcode.get(b_val_str, {})
                                evaluation = row_info.get("counter_evaluations", {}).get(c_label, {})
                                campaign_price = evaluation.get("customer_price")
                                if campaign_price is None:
                                    max_f_num = as_number(max_f)
                                    if max_f_num is not None:
                                        discount_amount = pe_item["discount_amount"]
                                        total_discount = round2(max_f_num * (discount_amount / 100.0)) if pe_item["discount_type"] == "%" else discount_amount
                                        campaign_price = round2(max_f_num - total_discount)
                                if campaign_price is not None:
                                    ws_pe.cell(r, fiyat_idx_pe).value = float(campaign_price)
                                keep_rows.append(r)

                        if keep_rows:
                            safe_keep_rows(ws_pe, keep_rows)
                            file_name = plus_extra_export_filename(pe_item, idx)
                            out_name = os.path.join(run_output_dir, file_name)
                            shrink_data_validations(ws_pe)
                            wb_pe.save(out_name)
                            fix_xlsx_for_trendyol(out_name)
                            generated_files.append(os.path.join(timestamp_folder, file_name))
        except Exception as e:
            print("Plus Ek İndirim export error:", e)

    # 5. Process Karşılamalı Kampanya (Çoklu Dosya Desteği)
    if target_type in ['Hepsi', 'Karşılamalı Kampanya']:
        try:
            from input_files import load_counter_configs
            counter_configs = load_counter_configs(INPUT_MANIFEST)
            
            for idx, c_item in enumerate(counter_configs):
                c_path = c_item.get('stored_path') or c_item.get('path')
                if c_path and os.path.exists(c_path):
                    wb_kars = openpyxl.load_workbook(c_path)
                    ws_kars = wb_kars.active
                    header_kars = [ws_kars.cell(1, c).value for c in range(1, ws_kars.max_column + 1)]
                    b_idx_kars = header_kars.index('Barkod') + 1 if 'Barkod' in header_kars else None
                    fiyat_idx_kars = header_kars.index('Kampanyalı Satış Fiyatı') + 1 if 'Kampanyalı Satış Fiyatı' in header_kars else None
                    max_fiyat_idx_kars = header_kars.index('Maksimum Girebileceğin Fiyat') + 1 if 'Maksimum Girebileceğin Fiyat' in header_kars else None
                    
                    if b_idx_kars and fiyat_idx_kars and max_fiyat_idx_kars:
                        c_label = c_item.get('label') or f"Karşılamalı #{idx+1}"
                        keep_rows = []
                        for r in range(2, ws_kars.max_row + 1):
                            b_val = ws_kars.cell(r, b_idx_kars).value
                            if not b_val: continue
                            
                            b_val_str = str(b_val).strip()
                            main_sel, extra_sel = get_selection(b_val_str)
                            should_keep = (extra_sel == c_label) or (target_type == "Karşılamalı Kampanya" and extra_sel.startswith("Karşılamalı"))

                            if should_keep:
                                max_fiyat_val = ws_kars.cell(r, max_fiyat_idx_kars).value
                                row_info = row_by_barcode.get(b_val_str, {})
                                evaluation = row_info.get('counter_evaluations', {}).get(c_label, {})
                                campaign_price = evaluation.get('price') or max_fiyat_val
                                if campaign_price:
                                    ws_kars.cell(r, fiyat_idx_kars).value = float(campaign_price)
                                keep_rows.append(r)
                                
                        if keep_rows:
                            safe_keep_rows(ws_kars, keep_rows)
                            
                            def format_num_clean_kars(val):
                                if val is None or val == "": return None
                                try:
                                    n = float(val)
                                    return f"{int(n)}" if n.is_integer() else f"{n}"
                                except (ValueError, TypeError):
                                    return str(val).strip()

                            min_p = format_num_clean_kars(c_item.get('min_price'))
                            disc = format_num_clean_kars(c_item.get('discount_amount'))
                            tp = format_num_clean_kars(c_item.get('trendyol_percent'))
                            disc_type = c_item.get('discount_type') or c_item.get('discount_unit') or 'TL'

                            if not min_p or not disc:
                                m_pct = re.search(r'(\d+(?:[\.,]\d+)?)\s*TL\s*Üzeri\s*/\s*%\s*(\d+(?:[\.,]\d+)?)', c_label, re.IGNORECASE)
                                if not m_pct:
                                    m_pct = re.search(r'(\d+(?:[\.,]\d+)?)\s*TL\s*Üzeri\s*/\s*(\d+(?:[\.,]\d+)?)\s*%', c_label, re.IGNORECASE)
                                if m_pct:
                                    min_p = min_p or format_num_clean_kars(m_pct.group(1))
                                    disc = disc or format_num_clean_kars(m_pct.group(2))
                                    disc_type = '%'
                                else:
                                    m = re.search(r'(\d+(?:[\.,]\d+)?)\s*TL\s*Üzeri\s*/\s*(\d+(?:[\.,]\d+)?)\s*TL', c_label, re.IGNORECASE)
                                    if m:
                                        min_p = min_p or format_num_clean_kars(m.group(1))
                                        disc = disc or format_num_clean_kars(m.group(2))

                            if min_p and disc:
                                disc_part = f"%{disc}" if disc_type == '%' else f"{disc}_TL"
                                if tp and tp != '0':
                                    out_filename = f"{min_p}_TL_Uzeri_{disc_part}_Indirim_%{tp}_Trendyol_Karsilamali.xlsx"
                                else:
                                    out_filename = f"{min_p}_TL_Uzeri_{disc_part}_Indirim_Trendyol_Karsilamali.xlsx"
                            else:
                                tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
                                clean_l = c_label.translate(tr_map)
                                safe_l = re.sub(r'[^\w%]', '_', clean_l)
                                safe_l = re.sub(r'_+', '_', safe_l).strip('_')
                                out_filename = f"Karsilamali_{safe_l}.xlsx"

                            out_name = os.path.join(run_output_dir, out_filename)
                            shrink_data_validations(ws_kars)
                            wb_kars.save(out_name)
                            fix_xlsx_for_trendyol(out_name)
                            generated_files.append(os.path.join(timestamp_folder, out_filename))
        except Exception:
            return processing_error("Karşılamalı Kampanya dosyası")

    # 6. Process Kupon (Çoklu Dosya Desteği)
    if target_type in ['Hepsi', 'Kupon', 'Plus Kupon', 'Karşılamalı Kampanya']:
        try:
            from input_files import load_coupon_configs
            coupon_configs = load_coupon_configs(INPUT_MANIFEST)
            
            for idx, cp_item in enumerate(coupon_configs):
                cp_path = cp_item.get('stored_path') or cp_item.get('path')
                if cp_path and os.path.exists(cp_path):
                    wb_cp = openpyxl.load_workbook(cp_path)
                    ws_cp = wb_cp.active
                    header_cp = [ws_cp.cell(1, c).value for c in range(1, ws_cp.max_column + 1)]
                    b_idx_cp = header_cp.index('Barkod') + 1 if 'Barkod' in header_cp else None
                    secim_idx_cp = header_cp.index('Eklenecek Ürünleri Seçiniz') + 1 if 'Eklenecek Ürünleri Seçiniz' in header_cp else None
                    
                    if b_idx_cp and secim_idx_cp:
                        cp_label = cp_item.get('label') or f"Kupon #{idx+1}"
                        keep_rows = []
                        for r in range(2, ws_cp.max_row + 1):
                            b_val = ws_cp.cell(r, b_idx_cp).value
                            if not b_val: continue
                            
                            b_val_str = str(b_val).strip()
                            main_sel, extra_sel = get_selection(b_val_str)
                            should_keep = (extra_sel == cp_label) or (target_type in ["Kupon", "Plus Kupon"] and extra_sel.startswith(cp_label.split()[0]))

                            if should_keep:
                                ws_cp.cell(r, secim_idx_cp).value = "Seçildi"
                                keep_rows.append(r)
                                
                        if keep_rows:
                            safe_keep_rows(ws_cp, keep_rows)
                            
                            def format_num_clean_cp(val):
                                if val is None or val == "": return None
                                try:
                                    n = float(val)
                                    return f"{int(n)}" if n.is_integer() else f"{n}"
                                except (ValueError, TypeError):
                                    return str(val).strip()

                            min_p = format_num_clean_cp(cp_item.get('min_price'))
                            disc = format_num_clean_cp(cp_item.get('discount_amount'))
                            tp = format_num_clean_cp(cp_item.get('trendyol_percent'))
                            disc_type = cp_item.get('discount_type') or 'TL'

                            if min_p and disc:
                                disc_part = f"%{disc}" if disc_type == '%' else f"{disc}_TL"
                                if tp and tp != '0':
                                    out_filename = f"{min_p}_TL_Uzerine_{disc_part}_Kupon_%{tp}_Trendyol_Plus.xlsx"
                                else:
                                    out_filename = f"{min_p}_TL_Uzerine_{disc_part}_Kupon_Trendyol_Plus.xlsx"
                            else:
                                tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
                                clean_l = cp_label.translate(tr_map)
                                safe_l = re.sub(r'[^\w%]', '_', clean_l)
                                safe_l = re.sub(r'_+', '_', safe_l).strip('_')
                                out_filename = f"Kupon_{safe_l}.xlsx"

                            out_name = os.path.join(run_output_dir, out_filename)
                            shrink_data_validations(ws_cp)
                            wb_cp.save(out_name)
                            fix_xlsx_for_trendyol(out_name)
                            generated_files.append(os.path.join(timestamp_folder, out_filename))
        except Exception as e:
            print("Kupon export error:", e)
            return processing_error("Kupon Kampanyası dosyası")

    # 7. Process Net İndirim (İndirim Oluştur Şablonu)
    if target_type in ['Hepsi', 'Net İndirim']:
        try:
            from input_files import load_net_discount_config
            nd_config = load_net_discount_config(INPUT_MANIFEST) or {}
            nd_label = nd_config.get('label') or build_campaign_label(nd_config, "net_discount")

            current_dict = {}
            if input_files.get("current") and os.path.exists(input_files["current"]):
                try:
                    c_df = pd.read_excel(input_files["current"])
                    if "Barkod" in c_df.columns:
                        c_df = c_df.assign(BARKOD_CLN=c_df["Barkod"].astype(str).str.strip())
                        current_dict = c_df.drop_duplicates(subset=["BARKOD_CLN"]).set_index("BARKOD_CLN").to_dict("index")
                except Exception:
                    pass

            selected_nd_barcodes = []
            for row_item in table_data:
                b_str = str(row_item.get('Barkod', '')).strip()
                if not b_str:
                    continue
                main_sel, extra_sel = get_selection(b_str)
                if extra_sel and (extra_sel == nd_label or extra_sel.endswith('Net İndirim')):
                    selected_nd_barcodes.append(b_str)

            if selected_nd_barcodes:
                nd_path = nd_config.get('stored_path') or nd_config.get('path') or input_files.get('net_discount')
                if nd_path and os.path.exists(nd_path):
                    wb_nd = openpyxl.load_workbook(nd_path)
                    ws_nd = wb_nd['Ürünler'] if 'Ürünler' in wb_nd.sheetnames else wb_nd.active
                    header_nd = [ws_nd.cell(1, c).value for c in range(1, ws_nd.max_column + 1)]
                    b_idx_nd = header_index(ws_nd, 'Barkod')
                    dahil_idx_nd = None
                    for c_idx, h_val in enumerate(header_nd, 1):
                        if h_val and 'dahil' in str(h_val).lower():
                            dahil_idx_nd = c_idx
                            break
                    if not dahil_idx_nd:
                        dahil_idx_nd = ensure_header(ws_nd, 'Kampayaya Dahil Edilsin Mi?')

                    id_idx_nd = header_index(ws_nd, 'Trendyol Ürün ID')
                    info_idx_nd = header_index(ws_nd, 'Ürün Bilgisi')
                    brand_idx_nd = header_index(ws_nd, 'Marka')
                    color_idx_nd = header_index(ws_nd, 'Renk')
                    model_idx_nd = header_index(ws_nd, 'Model Kodu')
                    price_idx_nd = header_index(ws_nd, 'Güncel Satış Fiyatı')
                    buybox_idx_nd = header_index(ws_nd, 'Buybox')

                    existing_barcodes = {}
                    for r in range(2, ws_nd.max_row + 1):
                        b_val = ws_nd.cell(r, b_idx_nd).value
                        if b_val:
                            existing_barcodes[str(b_val).strip()] = r

                    keep_rows = []
                    for b_str in selected_nd_barcodes:
                        if b_str in existing_barcodes:
                            r = existing_barcodes[b_str]
                            ws_nd.cell(r, dahil_idx_nd).value = "Evet"
                            keep_rows.append(r)
                        else:
                            r_new = ws_nd.max_row + 1
                            info = current_dict.get(b_str, {})
                            row_info = row_by_barcode.get(b_str, {})

                            link = str(info.get('Trendyol.com Linki') or '')
                            m_pid = re.search(r'-p-(\d+)', link)
                            pid = int(m_pid.group(1)) if m_pid else (info.get('Partner ID') or '')

                            if id_idx_nd: ws_nd.cell(r_new, id_idx_nd).value = pid
                            if info_idx_nd: ws_nd.cell(r_new, info_idx_nd).value = info.get('Ürün Adı') or row_info.get('Ürün Bilgisi') or ''
                            if brand_idx_nd: ws_nd.cell(r_new, brand_idx_nd).value = info.get('Marka') or row_info.get('Marka') or 'Paspas Yap'
                            if color_idx_nd: ws_nd.cell(r_new, color_idx_nd).value = info.get('Ürün Rengi') or row_info.get('Renk') or ''
                            ws_nd.cell(r_new, b_idx_nd).value = b_str
                            if model_idx_nd: ws_nd.cell(r_new, model_idx_nd).value = info.get('Model Kodu') or row_info.get('Model Kodu') or ''

                            price_val = info.get("Trendyol'da Satılacak Fiyat (KDV Dahil)") or info.get('Piyasa Satış Fiyatı (KDV Dahil)') or row_info.get('Güncel Ürün Fiyatı (TL)')
                            if price_idx_nd and price_val:
                                try:
                                    n_p = float(price_val)
                                    ws_nd.cell(r_new, price_idx_nd).value = f"{int(n_p)} ₺" if n_p.is_integer() else f"{n_p} ₺"
                                except Exception:
                                    ws_nd.cell(r_new, price_idx_nd).value = str(price_val)

                            if buybox_idx_nd: ws_nd.cell(r_new, buybox_idx_nd).value = info.get('Buybox') or 'Kaybeden'
                            ws_nd.cell(r_new, dahil_idx_nd).value = "Evet"
                            keep_rows.append(r_new)

                    if keep_rows:
                        safe_keep_rows(ws_nd, keep_rows)
                else:
                    wb_nd = openpyxl.Workbook()
                    ws_nd = wb_nd.active
                    ws_nd.title = "Ürünler"
                    headers = ['Trendyol Ürün ID', 'Ürün Bilgisi', 'Marka', 'Renk', 'Barkod', 'Model Kodu', 'Güncel Satış Fiyatı', 'Buybox', 'Kampayaya Dahil Edilsin Mi?']
                    ws_nd.append(headers)
                    for b_str in selected_nd_barcodes:
                        info = current_dict.get(b_str, {})
                        row_info = row_by_barcode.get(b_str, {})
                        link = str(info.get('Trendyol.com Linki') or '')
                        m_pid = re.search(r'-p-(\d+)', link)
                        pid = int(m_pid.group(1)) if m_pid else (info.get('Partner ID') or '')
                        price_val = info.get("Trendyol'da Satılacak Fiyat (KDV Dahil)") or info.get('Piyasa Satış Fiyatı (KDV Dahil)') or row_info.get('Güncel Ürün Fiyatı (TL)') or ''
                        price_str = ''
                        if price_val:
                            try:
                                n_p = float(price_val)
                                price_str = f"{int(n_p)} ₺" if n_p.is_integer() else f"{n_p} ₺"
                            except Exception:
                                price_str = str(price_val)
                        ws_nd.append([
                            pid,
                            info.get('Ürün Adı') or row_info.get('Ürün Bilgisi') or '',
                            info.get('Marka') or row_info.get('Marka') or 'Paspas Yap',
                            info.get('Ürün Rengi') or row_info.get('Renk') or '',
                            b_str,
                            info.get('Model Kodu') or row_info.get('Model Kodu') or '',
                            price_str,
                            info.get('Buybox') or 'Kaybeden',
                            'Evet'
                        ])

                def format_num_clean_nd(val):
                    if val is None or val == "": return None
                    try:
                        n = float(val)
                        return f"{int(n)}" if n.is_integer() else f"{n}"
                    except (ValueError, TypeError):
                        return str(val).strip()

                disc = format_num_clean_nd(nd_config.get('discount_amount'))
                disc_type = nd_config.get('discount_type', '%')
                if not disc and selected_nd_barcodes:
                    first_extra = get_selection(selected_nd_barcodes[0])[1]
                    if first_extra.startswith('%'):
                        disc = first_extra.replace('%', '').replace('Net İndirim', '').strip()
                        disc_type = '%'
                    elif 'TL' in first_extra:
                        disc = first_extra.replace('TL', '').replace('Net İndirim', '').strip()
                        disc_type = 'TL'

                disc_part = f"%{disc}" if disc_type == '%' else f"{disc}_TL"
                out_filename = f"Net_Indirim_{disc_part}_Urunler.xlsx" if disc else "Net_Indirim_Urunler.xlsx"
                out_name = os.path.join(run_output_dir, out_filename)
                shrink_data_validations(ws_nd)
                wb_nd.save(out_name)
                fix_xlsx_for_trendyol(out_name)
                generated_files.append(os.path.join(timestamp_folder, out_filename))
        except Exception as e:
            print("Net İndirim export error:", e)
            return processing_error("Net İndirim dosyası")

    # Birleşik Rapor Excel Çıktısı (Tek Dosya, 3 Sayfa: Genel Analiz, Özet Rapor, Uygulanmayan Ürünler)
    try:
        df_all = pd.DataFrame(table_data)
        if not df_all.empty:
            if 'checked' in df_all.columns:
                df_all = df_all.drop(columns=['checked'])
            # Seçimleri get_selection ile senkronize et
            df_all['Uygulanan Kampanya Seçimi'] = df_all['Barkod'].astype(str).str.strip().map(lambda b: get_selection(b)[0])
            df_all['Uygulanan Ekstra Kampanya Seçimi'] = df_all['Barkod'].astype(str).str.strip().map(lambda b: get_selection(b)[1])

            # Temiz Kolon Seti (Gereksiz, boş ve ham JSON/sözlük sütunları ayıklanmış)
            CLEAN_GENEL_COLUMNS = [
                'Barkod', 'Stok Adedi',
                'Uygulanan Kampanya Seçimi', 'Uygulanan Ekstra Kampanya Seçimi',
                'Önerilen Kampanya', 'Önerilen Ekstra Kampanya', 'Uygulanabilir Kampanyalar', 'İndirim Uygulanabilir',
                'Düşülebilecek Dip Fiyat (TL)', 'Uygulanabilecek İndirim (TL)', 'Uygulanabilecek İndirim (%)',
                'Mevcut İndirim (TL)', 'Mevcut İndirim (%)',
                'Güncel Ürün Fiyatı (TL)', 'Güncel Ürün Komisyon (%)', 'Güncel Ürün Kalan Net (TL)',
                'Avantajlı Ürün Fiyatı (YENİ TSF) (TL)', 'Avantajlı Ürün Komisyon (%)', 'Avantajlı Ürün Kalan Net (TL)',
                'Flaş Ürün 24 Saat Fiyatı (TL)', 'Flaş Ürün Komisyon (%)', 'Flaş Ürün Kalan Net (TL)',
                'Plus Fiyatı (TL)', 'Plus Tarife Seçimi', 'Plus Komisyon (%)', 'Plus Net (TL)',
                'Komisyon Tarifesi Fiyatı (TL)', 'Komisyon Tarifesi Komisyon (%)', 'Komisyon Tarifesi Net (TL)',
                'Komisyon Tarifesi Kademe', 'Komisyon Tarifesi Seçimi'
            ]
            present_genel_cols = [c for c in CLEAN_GENEL_COLUMNS if c in df_all.columns]
            df_genel_clean = df_all[present_genel_cols].copy()

            # 2. Sayfa: Özet Rapor (Sayfada seçilen görünür sütunlar)
            df_summary = build_report_dataframe(table_data, visible_columns)

            # 3. Sayfa: Uygulanmayan Ürünler
            UNAPPLIED_COLUMNS = [
                'Barkod', 'Stok Adedi',
                'Güncel Ürün Fiyatı (TL)', 'Güncel Ürün Komisyon (%)', 'Güncel Ürün Kalan Net (TL)',
                'Düşülebilecek Dip Fiyat (TL)', 'Uygulanabilecek İndirim (TL)', 'Uygulanabilecek İndirim (%)',
                'Önerilen Kampanya', 'Uygulanabilir Kampanyalar'
            ]
            df_unapp = df_all[
                (df_all['Uygulanan Kampanya Seçimi'] == 'Hiçbiri')
                & (df_all['Uygulanan Ekstra Kampanya Seçimi'] == 'Hiçbiri')
            ].copy()
            unapp_present = [c for c in UNAPPLIED_COLUMNS if c in df_unapp.columns]
            df_unapplied_clean = df_unapp[unapp_present].copy()

            # Tek Birleşik Excel Dosyası Olarak Kaydet
            out_report = os.path.join(run_output_dir, "Kampanya_Genel_Raporu.xlsx")
            with pd.ExcelWriter(out_report, engine="openpyxl") as writer:
                df_genel_clean.to_excel(writer, sheet_name="Genel_Analiz_Raporu", index=False)
                df_summary.to_excel(writer, sheet_name="Kampanya_Ozet_Raporu", index=False)
                df_unapplied_clean.to_excel(writer, sheet_name="Uygulanmayan_Urunler", index=False)

            # Formatlama (A2 Sabitleme, Otomatik Filtre ve Kolon Genişlikleri)
            wb_rep = openpyxl.load_workbook(out_report)
            for sheet in wb_rep.worksheets:
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                for col_idx in range(1, sheet.max_column + 1):
                    letter = openpyxl.utils.get_column_letter(col_idx)
                    max_len = max(len(str(sheet.cell(1, col_idx).value or "")), 10)
                    sheet.column_dimensions[letter].width = max(max_len + 3, 12)
            wb_rep.save(out_report)
            fix_xlsx_for_trendyol(out_report)
            generated_files.append(os.path.join(timestamp_folder, "Kampanya_Genel_Raporu.xlsx"))
    except Exception as e:
        print("Rapor oluşturma hatası:", e)
        return processing_error("Rapor")
        
    files_str = "\n".join([f"- {f}" for f in generated_files])
    return jsonify({
        "success": True, 
        "message": f"Tarihlere göre {len(generated_files)} dosya başarıyla oluşturuldu!\n\nOluşturulan Dosyalar:\n{files_str}",
        "generated_files": generated_files,
        "timestamp_folder": timestamp_folder
    })

if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", 5114))
    print(f"\n=======================================================")
    print(f"🚀 Trendyol Kampanya Hesaplama Motoru Aktif")
    print(f"📍 Localhost / Yerel Tarayıcı : http://localhost:{port}")
    print(f"📍 Yerel IP                   : http://127.0.0.1:{port}")
    print(f"🌐 Ağ / Tüm Arayüzler (Host)  : http://0.0.0.0:{port}")
    print(f"🔗 Ngrok ile Dışarı Açma      : ngrok http {port}")
    print(f"=======================================================\n")
    app.run(host=host, port=port, debug=True)
